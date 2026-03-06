#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
VRP Optimizer Pro v4 - Serveur Flask avec Groq API
Normalisation adresses  Geocodage intelligent  Routing IA  Dashboard
"""

from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
import os, math, base64, re, time, logging, threading
from io import BytesIO
from datetime import datetime, timedelta
import json
import unicodedata
from dotenv import load_dotenv

load_dotenv()

import pandas as pd
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from werkzeug.utils import secure_filename

from ortools.constraint_solver import pywrapcp, routing_enums_pb2
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────
# FLASK SETUP
# ─────────────────────────────────────────────────────────────────
app = Flask(__name__)
CORS(app)

UPLOAD_FOLDER = '/tmp/vrp_uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 32 * 1024 * 1024

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s  %(levelname)-8s  %(message)s',
    datefmt='%H:%M:%S',
)
log = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────
# CONSTANTES & CONFIG
# ─────────────────────────────────────────────────────────────────
BAN_API_URL   = "https://api-adresse.data.gouv.fr/search/"
NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"
OSRM_TABLE    = "https://router.project-osrm.org/table/v1"
GEOCODE_DELAY = 1.1
CHUNK_SIZE    = 50
USER_AGENT    = "VRPOptimizer/4.0 contact@monentreprise.com"
GROQ_API_URL  = "https://api.groq.com/openai/v1/chat/completions"
GROQ_API_KEY  = os.environ.get("GROQ_API_KEY", "")
# Models gratuits Groq: llama-3.3-70b-versatile, gemma2-9b-it, mixtral-8x7b-32768
GROQ_MODEL    = os.environ.get("GROQ_MODEL", "llama-3.3-70b-versatile")

# ─────────────────────────────────────────────────────────────────
# LEXIQUE DE NORMALISATION DES ADRESSES
# ─────────────────────────────────────────────────────────────────
ADDRESS_NORMALIZATION = {
    # Types de voies
    'bd': 'boulevard', 'b.d': 'boulevard', 'bld': 'boulevard',
    'av': 'avenue', 'a.v': 'avenue', 'ave': 'avenue',
    'rue': 'rue', 'r': 'rue',
    'pl': 'place', 'pl.': 'place',
    'sq': 'square', 'sq.': 'square',
    'imp': 'impasse', 'imp.': 'impasse',
    'all': 'allee', 'all.': 'allee',
    'crs': 'cours', 'cours': 'cours',
    'ch': 'chemin', 'ch.': 'chemin',
    'ctl': 'casterale', 'ctle': 'casterale',
    'quai': 'quai', 'q': 'quai',
    'bdg': 'boulevard', 'boul': 'boulevard', 'boulv': 'boulevard',
    'faub': 'faubourg', 'fg': 'faubourg',
    'rocade': 'rocade', 'rce': 'rocade',
    'route': 'route', 'rte': 'route',
    'bretelle': 'bretelle', 'brt': 'bretelle',
    'passage': 'passage', 'pass': 'passage',
    ' traverse': 'traverse', 'trav': 'traverse',
    'vil': 'villa', 'villas': 'villa',
    'cite': 'cite', 'cite': 'cite', 'cte': 'cite',
    'res': 'residence', 'res': 'residence', 'residence': 'residence',
    'lot': 'lotissement', 'lot.': 'lotissement',
    'zac': 'zone d\'activite commerciale', 'z.i': 'zone industrielle',
    'za': 'zone d\'activite', 'zad': 'zone d\'amenagement differe',
    
    # Prefixes numeriques
    'bis': 'bis', 'ter': 'ter', 'quater': 'quater',
    'b': 'bis', 't': 'ter', 'q': 'quater',
    
    # Autres abreviations courantes
    'st': 'saint', 'ste': 'sainte',
    'mle': 'maison', 'maison': 'maison',
    'bat': 'batiment', 'bat': 'batiment', 'batiment': 'batiment',
    'etg': 'etage', 'etage': 'etage',
    'apt': 'appartement', 'app': 'appartement',
    'bp': 'boite postale', 'b.p': 'boite postale',
    'cedex': 'cedex', 'cx': 'cedex',
    'cs': 'coursier', 'c.s': 'coursier',
    'ts': 'terrasse', 'tce': 'terrasse',
    
    # Nettoyage
    ',': ' ', ';': ' ', '/': ' ', '-': ' ', '_': ' ',
    '.': ' ', '  ': ' ', '   ': ' ',
}

# Codes postaux francais pour validation
POSTAL_CODES_PATTERN = re.compile(r'\b(0[1-9]|[1-8]\d|9[0-5])\d{3}\b')

# ─────────────────────────────────────────────────────────────────
# SESSION HTTP
# ─────────────────────────────────────────────────────────────────
def _make_session():
    s = requests.Session()
    s.headers.update({"User-Agent": USER_AGENT})
    retry = Retry(total=2, backoff_factor=1, status_forcelist=[500, 502, 503, 504])
    s.mount("https://", HTTPAdapter(max_retries=retry))
    s.mount("http://", HTTPAdapter(max_retries=retry))  # ✅ Ajouter HTTP aussi !
    return s

SESSION = _make_session()

# ─────────────────────────────────────────────────────────────────
# STOCKAGE TEMPORAIRE (pour la validation manuelle)
# ─────────────────────────────────────────────────────────────────
temp_data_store = {}

# ─────────────────────────────────────────────────────────────────
# PROGRESSION TEMPS-REEL (SSE)
# ─────────────────────────────────────────────────────────────────
progress_store: dict = {}
progress_lock  = threading.Lock()

def set_progress(pid: str, pct: int, msg: str = "", done: bool = False):
    """Met a jour la progression pour un progress_id donne."""
    if not pid:
        return
    with progress_lock:
        progress_store[pid] = {"pct": pct, "msg": msg, "done": done}

@app.route("/api/progress/<progress_id>")
def api_progress_stream(progress_id):
    """SSE : streame la progression d'une operation longue."""
    def generate():
        last_pct = -1
        max_wait = 300  # 5 min max
        waited   = 0
        while waited < max_wait:
            with progress_lock:
                data = progress_store.get(progress_id, {})
            pct  = data.get("pct", 0)
            msg  = data.get("msg", "")
            done = data.get("done", False)
            if pct != last_pct or done:
                yield f"data: {json.dumps({'pct': pct, 'msg': msg, 'done': done})}\n\n"
                last_pct = pct
            if done:
                with progress_lock:
                    progress_store.pop(progress_id, None)
                break
            time.sleep(0.25)
            waited += 0.25
    return app.response_class(
        generate(),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"}
    )

# ─────────────────────────────────────────────────────────────────
# AUTO-DETECTION ET MAPPAGE DES COLONNES
# ─────────────────────────────────────────────────────────────────
FIELD_META = {
    'site':       {'label': 'Nom du site',          'required': True,  'keywords': ['site','nom','name','client','lieu','point','etablissement','enseigne','magasin','destinataire','libelle','description']},
    'address':    {'label': 'Adresse',               'required': True,  'keywords': ['adresse','address','rue','voie','street','localisation','domicile','voirie','adr','adres']},
    'ville':      {'label': 'Ville / Code postal',   'required': True,  'keywords': ['ville','city','commune','localite','municipality','cp','codepostal','code_postal','code postal','agglomeration']},
    'date_min':   {'label': 'Date debut',            'required': False, 'keywords': ['date_min','debut','from','start','d_min','date_debut','date_from','ouverture','datemin']},
    'date_max':   {'label': 'Date fin',              'required': False, 'keywords': ['date_max','fin','to','end','d_max','date_fin','date_to','cloture','datemax']},
    'temps_site': {'label': 'Temps de visite (min)', 'required': False, 'keywords': ['temps','duree','duration','visite','minutes','heures','tps','temps_site','time_on_site','time']},
}

def _norm_col(s: str) -> str:
    """Normalise un nom de colonne pour la comparaison."""
    s = s.lower().strip()
    for acc, plain in [('e','e'),('e','e'),('e','e'),('e','e'),('a','a'),('a','a'),
                       ('u','u'),('u','u'),('i','i'),('i','i'),('o','o'),('c','c')]:
        s = s.replace(acc, plain)
    return re.sub(r'[^a-z0-9]', '_', s)

def auto_detect_columns(columns: list) -> dict:
    """Retourne {field: colonne_source | None} par correspondance de mots-cles."""
    norm = {col: _norm_col(col) for col in columns}
    result = {}
    for field, info in FIELD_META.items():
        match = None
        # 1. Correspondance exacte sur le nom de champ
        for col, nc in norm.items():
            if nc == field:
                match = col; break
        if not match:
            # 2. Un mot-cle contenu dans le nom de colonne (ou l'inverse)
            for col, nc in norm.items():
                for kw in info['keywords']:
                    nkw = _norm_col(kw)
                    if nkw in nc or nc in nkw:
                        match = col; break
                if match:
                    break
        result[field] = match
    return result

def apply_column_mapping(df, mapping: dict):
    """Renomme les colonnes du DataFrame selon le mapping {field: source_col}."""
    rename = {}
    for field, src in mapping.items():
        if src and src in df.columns and src != field:
            rename[src] = field
    return df.rename(columns=rename) if rename else df

@app.route("/api/preview-columns", methods=["POST"])
def api_preview_columns():
    """Analyse un fichier Excel : retourne colonnes, suggestions de mappage et apercu."""
    try:
        if "file" not in request.files:
            return jsonify({"success": False, "error": "Aucun fichier"}), 400
        raw = request.files["file"].read()
        df  = pd.read_excel(BytesIO(raw), dtype=str, na_filter=False, keep_default_na=False, nrows=10)
        if df.empty:
            return jsonify({"success": False, "error": "Fichier vide"}), 400
        columns     = list(df.columns)
        suggestions = auto_detect_columns(columns)
        preview     = df.head(5).fillna("").to_dict("records")
        fields_meta = {k: {"label": v["label"], "required": v["required"]} for k, v in FIELD_META.items()}
        return jsonify({"success": True, "columns": columns,
                        "suggestions": suggestions, "preview": preview,
                        "fields_meta": fields_meta})
    except Exception as e:
        log.exception("Erreur /api/preview-columns")
        return jsonify({"success": False, "error": str(e)}), 500

# ─────────────────────────────────────────────────────────────────
# NORMALISATION D'ADRESSES
# ─────────────────────────────────────────────────────────────────
def normalize_address(address, ville):
    """
    Normalise une adresse en utilisant le lexique
    Retourne l'adresse normalisee et les corrections appliquees
    """
    if pd.isna(address) or not address:
        return "", []
    
    original = str(address).strip()
    corrections = []
    
    # Etape 1 : Nettoyage initial (caracteres speciaux, espaces multiples)
    original = re.sub(r'[;,/_]', ' ', original)
    original = re.sub(r'\s+', ' ', original).strip()
    
    # Etape 2 : Remplacements intelligents par ordre de priorite
    # On utilise des regex avec word boundaries pour eviter les remplacements partiels
    
    replacements = [
        # Types de voies (avec word boundaries pour eviter les remplacements dans les mots)
        (r'\bBD\b', 'boulevard'),
        (r'\bB\.D\.', 'boulevard'),
        (r'\bBLD\b', 'boulevard'),
        (r'\bAV\b', 'avenue'),
        (r'\bA\.V\.', 'avenue'),
        (r'\bAVE\b', 'avenue'),
        (r'\bPL\b', 'place'),
        (r'\bPL\.', 'place'),
        (r'\bSQ\b', 'square'),
        (r'\bSQ\.', 'square'),
        (r'\bIMP\b', 'impasse'),
        (r'\bIMP\.', 'impasse'),
        (r'\bALL\b', 'allee'),
        (r'\bALL\.', 'allee'),
        (r'\bCRS\b', 'cours'),
        (r'\bCH\b', 'chemin'),
        (r'\bCH\.', 'chemin'),
        (r'\bCTL\b', 'casterale'),
        (r'\bQUAI\b', 'quai'),
        (r'\bBDG\b', 'boulevard'),
        (r'\bBOUL\b', 'boulevard'),
        (r'\bBOULV\b', 'boulevard'),
        (r'\bFAUB\b', 'faubourg'),
        (r'\bFG\b', 'faubourg'),
        (r'\bRTE\b', 'route'),
        (r'\bBRTE\b', 'bretelle'),
        (r'\bPASS\b', 'passage'),
        (r'\bPASS\.', 'passage'),
        (r'\bTRAV\b', 'traverse'),
        (r'\bVIL\b', 'villa'),
        (r'\bCITE\b', 'cite'),
        (r'\bCTE\b', 'cite'),
        (r'\bRES\b', 'residence'),
        (r'\bLOT\b', 'lotissement'),
        (r'\bZAC\b', 'zone d\'activite commerciale'),
        (r'\bZ\.I\.', 'zone industrielle'),
        (r'\bZA\b', 'zone d\'activite'),
        
        # Prefixes numeriques
        (r'\bBIS\b', 'bis'),
        (r'\bTER\b', 'ter'),
        (r'\bQUATER\b', 'quater'),
        
        # Saints/Saintes
        (r'\bST\s', 'saint '),
        (r'\bSTE\s', 'sainte '),
        
        # Autres abreviations
        (r'\bMLE\b', 'maison'),
        (r'\bBAT\b', 'batiment'),
        (r'\bBAT\.', 'batiment'),
        (r'\bBT\b', 'batiment'),
        (r'\bETG\b', 'etage'),
        (r'\bAPT\b', 'appartement'),
        (r'\bAPP\b', 'appartement'),
        (r'\bBP\b', 'boite postale'),
        (r'\bB\.P\.', 'boite postale'),
        (r'\bCEDEX\b', 'cedex'),
        (r'\bCX\b', 'cedex'),
        (r'\bCS\b', 'coursier'),
        (r'\bC\.S\.', 'coursier'),
        (r'\bTCE\b', 'terrasse'),
        (r'\bTS\b', 'terrasse'),
    ]
    
    # Etape 3 : Appliquer les remplacements avec word boundaries
    normalized = original.lower()
    for pattern, replacement in replacements:
        old = normalized
        normalized = re.sub(pattern, replacement, normalized, flags=re.IGNORECASE)
        if old != normalized:
            corrections.append(f"{pattern}  {replacement}")
    
    # Etape 4 : Nettoyage final
    normalized = re.sub(r'\s+', ' ', normalized).strip()
    normalized = normalized.title()
    
    # Etape 5 : Ajout de la ville si pas presente
    if ville and pd.notna(ville):
        ville_clean = str(ville).strip().title()
        if ville_clean.lower() not in normalized.lower():
            normalized = f"{normalized}, {ville_clean}"
    
    return normalized, list(set(corrections))

def validate_address_format(address):
    """
    Valide le format d'une adresse
    Retourne (valid, errors, suggestions)
    """
    if not address or len(address) < 5:
        return False, ["Adresse trop courte"], []
    
    errors = []
    suggestions = []
    
    # Verifier presence d'un numero de rue
    if not re.search(r'\b\d+\b', address):
        errors.append("Pas de numero de rue detecte")
        suggestions.append("Ajoutez un numero si possible")
    
    # Verifier presence d'un type de voie
    voie_types = ['rue', 'avenue', 'boulevard', 'place', 'chemin', 'impasse', 
                  'allee', 'cours', 'route', 'square', 'quai', 'residence']
    has_voie = any(vt in address.lower() for vt in voie_types)
    if not has_voie:
        errors.append("Type de voie non detecte")
        suggestions.append("Verifiez le type de voie (rue, avenue, etc.)")
    
    # Verifier code postal
    has_cp = POSTAL_CODES_PATTERN.search(address) is not None
    if not has_cp:
        errors.append("Code postal non detecte")
        suggestions.append("Ajoutez le code postal pour ameliorer la precision")
    
    return len(errors) == 0, errors, suggestions

# ─────────────────────────────────────────────────────────────────
# GEOCODAGE ROBUSTE - TOUJOURS TROUVER UN RESULTAT
# ─────────────────────────────────────────────────────────────────

# Centres approximatifs des departements francais (code postal -> lat, lon)
DEPT_CENTERS = {
    '01': (46.18, 5.45), '02': (49.55, 3.62), '03': (46.33, 3.25), '04': (44.05, 6.12), '05': (44.55, 6.35),
    '06': (43.70, 7.26), '07': (44.83, 4.50), '08': (49.53, 4.75), '09': (43.05, 1.55), '10': (48.30, 4.36),
    '11': (43.20, 2.87), '12': (44.55, 2.70), '13': (43.52, 5.20), '14': (49.03, -0.50), '15': (45.06, 2.50),
    '16': (45.65, 0.16), '17': (45.74, -0.65), '18': (47.50, 2.40), '19': (45.26, 1.77), '20': (42.03, 9.14),
    '21': (47.32, 4.80), '22': (48.52, -2.76), '23': (45.84, 1.94), '24': (45.03, 0.72), '25': (47.32, 6.32),
    '26': (44.72, 5.06), '27': (49.20, 1.05), '28': (48.44, 1.50), '29': (48.20, -4.03), '30': (43.83, 4.35),
    '31': (43.20, 1.44), '32': (43.51, 0.71), '33': (44.84, -0.58), '34': (43.61, 3.52), '35': (48.20, -1.67),
    '36': (46.62, 1.51), '37': (47.21, 0.69), '38': (45.30, 5.56), '39': (46.69, 5.79), '40': (43.97, -0.77),
    '41': (47.72, 1.25), '42': (45.52, 4.21), '43': (45.20, 3.74), '44': (47.22, -1.55), '45': (47.84, 2.00),
    '46': (44.52, 1.76), '47': (44.52, 0.71), '48': (44.52, 3.52), '49': (47.50, -0.55), '50': (49.02, -1.50),
    '51': (49.02, 4.37), '52': (48.30, 5.00), '53': (48.01, -0.55), '54': (48.69, 6.18), '55': (49.02, 5.20),
    '56': (47.76, -2.80), '57': (49.00, 6.35), '58': (47.01, 3.50), '59': (50.63, 3.20), '60': (49.44, 2.45),
    '61': (48.52, 0.20), '62': (50.52, 2.45), '63': (45.78, 3.08), '64': (43.30, -0.58), '65': (43.04, 0.21),
    '66': (42.50, 2.52), '67': (48.52, 7.56), '68': (47.84, 7.24), '69': (45.76, 4.83), '70': (47.53, 6.00),
    '71': (46.69, 4.34), '72': (48.00, 0.20), '73': (45.55, 6.02), '74': (46.01, 6.32), '75': (48.85, 2.35),
    '76': (49.50, 1.00), '77': (48.60, 3.00), '78': (48.80, 1.80), '79': (46.33, -0.53), '80': (50.00, 2.30),
    '81': (43.84, 2.03), '82': (44.23, 1.50), '83': (43.43, 6.24), '84': (44.00, 5.54), '85': (46.69, -1.50),
    '86': (46.52, 0.53), '87': (45.84, 1.33), '88': (48.20, 6.52), '89': (47.80, 3.70), '90': (47.63, 6.99),
    '91': (48.60, 2.33), '92': (48.84, 2.20), '93': (48.90, 2.41), '94': (48.80, 2.51), '95': (49.10, 2.10),
    '971': (16.24, -61.54), '972': (14.61, -61.00), '973': (4.93, -53.23), '974': (-21.11, 55.53), '976': (-12.83, 45.23)
}

# Cache geocodage en memoire (cle = adresse normalisee, valeur = resultat)
_geocode_cache: dict = {}

def _try_geocode_ban(address, timeout=10):
    """
    Geocode via API BAN (Base Adresse Nationale).
    Gratuit, sans rate limit, optimise pour les adresses francaises.
    """
    try:
        r = SESSION.get(BAN_API_URL, params={
            "q": address,
            "limit": 1,
        }, timeout=timeout)
        r.raise_for_status()
        data = r.json()
        features = data.get("features", [])
        if features:
            coords = features[0]["geometry"]["coordinates"]
            props = features[0]["properties"]
            score = props.get("score", 0)
            if score < 0.3:
                return None
            return {
                "lat": coords[1],
                "lon": coords[0],
                "display_name": props.get("label", ""),
                "score": score,
            }
        return None
    except Exception:
        return None


def _try_geocode(address, timeout=5):
    """
    Tente de geocoder une adresse via Nominatim (fallback).
    """
    try:
        r = SESSION.get(NOMINATIM_URL, params={
            "q": address,
            "format": "json",
            "limit": 1,
            "countrycodes": "fr",
            "accept-language": "fr-FR"
        }, timeout=timeout)
        r.raise_for_status()
        data = r.json()

        if data and len(data) > 0:
            return {
                "lat": float(data[0]["lat"]),
                "lon": float(data[0]["lon"]),
                "display_name": data[0].get("display_name", "")
            }
        return None
    except Exception as e:
        return None

def _geocode_one(address, raw_address=None):
    """
    Geocode une adresse avec STRATEGIES MULTIPLES
    NE JAMAIS RETOURNER None - TOUJOURS TROUVER UN RESULTAT
    Utilise un cache en memoire pour eviter les requetes redondantes.
    """
    if not address:
        log.warning("  Adresse vide - Utilisation centre France")
        return {
            "lat": 46.603354,
            "lon": 1.888334,
            "display_name": "Centre de la France (adresse vide)",
            "is_fallback": True
        }
    
    address = str(address).strip()

    # Verifier le cache
    if address in _geocode_cache:
        return _geocode_cache[address]

    # Extraire les composants de l'adresse
    cp_match = re.search(r'\b(0[1-9]|[1-8]\d|9[0-5])\d{3}\b', address)
    code_postal = cp_match.group() if cp_match else None
    dept_code = code_postal[:2] if code_postal else None
    
    # Extraire la ville (apres le code postal)
    if code_postal:
        after_cp = address[cp_match.end():].strip(', ')
        ville = after_cp if after_cp else None
    else:
        parts = [p.strip() for p in address.split(',')]
        ville = parts[-1] if len(parts) > 1 else None
    
    # Strategies de recherche
    strategies = [
        raw_address,  # Strategie 0: Adresse brute originale (avant normalisation)
        address,  # Strategie 1: Adresse normalisee
        f"{code_postal} {ville}".strip() if code_postal and ville else None,  # Strategie 2: CP + ville
        ville if ville else None,  # Strategie 3: Juste ville
    ]

    # Filtrer les strategies None/vides
    strategies = [s for s in strategies if s and len(s) > 2]

    # --- API BAN (Base Adresse Nationale) : rapide, sans rate limit ---
    for i, strategy in enumerate(strategies):
        result = _try_geocode_ban(strategy)
        if result:
            log.info(f"✅ BAN trouve (strat {i}): {strategy[:40]}  {result['lat']:.4f}, {result['lon']:.4f}")
            _geocode_cache[address] = result
            return result

    # --- Fallback Nominatim (avec rate limit 1 req/s) ---
    for i, strategy in enumerate(strategies):
        result = _try_geocode(strategy, timeout=5)
        if result:
            log.info(f"✅ Nominatim trouve (strat {i}): {strategy[:40]}  {result['lat']:.4f}, {result['lon']:.4f}")
            _geocode_cache[address] = result
            return result
        if i < len(strategies) - 1:
            time.sleep(GEOCODE_DELAY)

    if dept_code and dept_code in DEPT_CENTERS:
        lat, lon = DEPT_CENTERS[dept_code]
        log.warning(f"  Fallback departement {dept_code}: {address[:40]}  {lat:.4f}, {lon:.4f}")
        result = {
            "lat": lat,
            "lon": lon,
            "display_name": f"Centre departement {dept_code} ({code_postal})",
            "is_fallback": True
        }
        _geocode_cache[address] = result
        return result

    # Ult recours: Centre de la France
    log.error(f"🚨 ULTIME RECOURS: {address[:40]}  Centre France")
    return {
        "lat": 46.603354,
        "lon": 1.888334,
        "display_name": "Centre de la France (dernier recours)",
        "is_fallback": True
    }

def geocode_with_validation(raw_rows, session_id, progress_callback=None):
    """
    Geocode les sites avec validation et retourne les resultats.
    progress_callback(pct, msg) est appele apres chaque adresse.
    """
    results = {
        "success": [],
        "failed": [],
        "warnings": [],
        "total": len(raw_rows),
        "geocoded": 0,
        "manual_needed": 0
    }

    log.info(" Geocodage de %d adresses avec validation", len(raw_rows))

    for i, row in enumerate(raw_rows):
        log.debug(f"[{i+1}/{len(raw_rows)}] Debut traitement adresse {i}")
        row_dict = dict(row)
        row_dict["_original_index"] = i  # IMPORTANT : sauvegarder l'index original

        # Fallback : si 'site' vide, utiliser d'autres colonnes candidates
        if not row_dict.get("site"):
            for _fallback in ("client", "nom", "name", "etablissement", "raison_sociale", "libelle"):
                if row_dict.get(_fallback):
                    row_dict["site"] = row_dict[_fallback]
                    break

        original_address = row_dict.get("address", "")
        ville = row_dict.get("ville", "")
        
        # Verifier coordonnees deja presentes
        try:
            lat = float(str(row_dict.get("lat", "")).replace(",", "."))
            lon = float(str(row_dict.get("lon", "")).replace(",", "."))
            if math.isfinite(lat) and math.isfinite(lon):
                row_dict["_lat"] = lat
                row_dict["_lon"] = lon
                row_dict["_normalized_address"] = original_address
                row_dict["_corrections"] = []
                results["success"].append(row_dict)
                results["geocoded"] += 1
                log.info("[%d/%d] ✅ Coordonnees directes", i+1, len(raw_rows))
                continue
        except (TypeError, ValueError):
            pass
        
        # Normaliser l'adresse
        log.debug(f"[{i+1}] Normalisation de: {original_address[:50]}")
        normalized, corrections = normalize_address(original_address, ville)
        row_dict["_normalized_address"] = normalized
        row_dict["_corrections"] = corrections
        
        # Valider le format
        log.debug(f"[{i+1}] Validation de: {normalized[:50]}")
        is_valid, errors, suggestions = validate_address_format(normalized)
        
        if not is_valid:
            results["warnings"].append({
                "index": i,
                "site": row_dict.get("site", ""),
                "original": original_address,
                "normalized": normalized,
                "errors": errors,
                "suggestions": suggestions
            })
        
        # Tenter le geocodage (le sleep Nominatim est saute si cache hit)
        log.debug(f"[{i+1}] Geocodage de: {normalized[:50]}")
        cache_hit = normalized in _geocode_cache
        # Construire la requete brute (sans normalisation) pour la strategie 0
        raw_query = original_address
        if ville and pd.notna(ville):
            ville_str = str(ville).strip()
            if ville_str.lower() not in original_address.lower():
                raw_query = f"{original_address}, {ville_str}"
        geo_result = _geocode_one(normalized, raw_address=raw_query if raw_query != normalized else None)
        
        log.debug(f"[{i+1}] Resultat: lat={geo_result.get('lat')}, lon={geo_result.get('lon')}, fallback={geo_result.get('is_fallback', False)}")

        # IMPORTANT: _geocode_one NE RETOURNE JAMAIS None
        if geo_result:
            row_dict["_lat"] = geo_result["lat"]
            row_dict["_lon"] = geo_result["lon"]
            row_dict["_display_name"] = geo_result["display_name"]
            row_dict["_is_fallback"] = geo_result.get("is_fallback", False)

            if geo_result.get("is_fallback"):
                results["warnings"].append({
                    "index": i,
                    "site": row_dict.get("site", ""),
                    "type": "fallback_geocode",
                    "message": f"Coordonnees approximatives: {geo_result['display_name']}",
                    "original": original_address,
                    "normalized": normalized
                })

            results["success"].append(row_dict)
            results["geocoded"] += 1
            log.info("[%d/%d] %s %s  %.5f, %.5f%s", i+1, len(raw_rows),
                    "⚡" if cache_hit else "✅",
                    normalized[:30], geo_result["lat"], geo_result["lon"],
                    " (fallback)" if geo_result.get("is_fallback") else "")
        else:
            log.error("[%d/%d]  ERREUR CRITIQUE: %s", i+1, len(raw_rows), normalized)
            row_dict["_lat"] = 46.603354
            row_dict["_lon"] = 1.888334
            row_dict["_display_name"] = "Centre France (erreur critique)"
            row_dict["_is_fallback"] = True
            results["success"].append(row_dict)
            results["geocoded"] += 1

        # Delai Nominatim uniquement si Nominatim a ete utilise (pas BAN, pas cache)
        if not cache_hit and not geo_result.get("score"):
            time.sleep(GEOCODE_DELAY)

        # Progression temps-reel
        if progress_callback:
            pct = int(10 + (i + 1) / len(raw_rows) * 80)
            progress_callback(pct, f"Geocodage {i+1}/{len(raw_rows)} : {normalized[:35]}")

    # Sauvegarder les resultats pour validation manuelle
    temp_data_store[session_id] = results
    
    return results


# ─────────────────────────────────────────────────────────────────
# INTEGRATION GROQ API
# ─────────────────────────────────────────────────────────────────
def _site_name_candidates(row):
    """Retourne les identifiants texte possibles d'un site (normalises)."""
    keys = (
        "site", "client", "nom", "name", "etablissement", "raison_sociale", "libelle", "description"
    )
    out = []
    for k in keys:
        v = row.get(k, "")
        s = str(v).strip()
        if s:
            out.append(s.lower())
    return out


def _norm_site_token(value):
    """Normalise un identifiant site/depot pour comparaison robuste."""
    s = str(value or "").strip().lower()
    if not s:
        return ""
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r"[^a-z0-9]+", "", s)
    return s


def _row_contains_depot_keyword(row):
    """Heuristique de secours: une ligne ressemble au depot."""
    for v in _site_name_candidates(row):
        nv = _norm_site_token(v)
        if "depot" in nv:
            return True
    return False


def _row_matches_depot(row, depot_name):
    depot_norm = _norm_site_token(depot_name)
    if not depot_norm:
        return False
    for v in _site_name_candidates(row):
        cand = _norm_site_token(v)
        if not cand:
            continue
        if cand == depot_norm:
            return True
        if len(cand) >= 8 and len(depot_norm) >= 8 and (cand in depot_norm or depot_norm in cand):
            return True
    return False


def _find_depot_index(rows, depot_name):
    """Trouve l'index du depot dans une liste de lignes."""
    for i, row in enumerate(rows):
        if _row_matches_depot(row, depot_name):
            return i

    fallback = [i for i, row in enumerate(rows) if _row_contains_depot_keyword(row)]
    if len(fallback) == 1:
        log.warning("   Depot retrouve via heuristique 'depot' (format de nom different)")
        return fallback[0]
    return None


def _visit_minutes_from_row(row, default=30):
    """Retourne le temps de visite en minutes depuis temps_site ou Temps (heures)."""
    raw_h = str(row.get("Temps", "")).strip()
    if raw_h:
        try:
            h = float(raw_h.replace(",", "."))
            return int(round(h * 60)) if h > 0 else default
        except (ValueError, TypeError):
            pass

    raw_min = str(row.get("temps_site", "")).strip()
    if raw_min:
        try:
            v = int(float(raw_min.replace(",", ".")))
            return v if v > 0 else default
        except (ValueError, TypeError):
            pass

    return default


def ask_groq(prompt, model=GROQ_MODEL, timeout=60):
    """
    Envoie un prompt a Groq API. Retourne la reponse texte ou None.
    """
    if not GROQ_API_KEY:
        log.warning("Cle GROQ_API_KEY non configuree")
        return None
    try:
        log.info("Envoi prompt Groq (timeout=%ds, model=%s, len=%d)", timeout, model, len(prompt))
        response = requests.post(
            GROQ_API_URL,
            headers={
                "Authorization": f"Bearer {GROQ_API_KEY}",
                "Content-Type": "application/json",
            },
            json={
                "model": model,
                "messages": [{"role": "user", "content": prompt}],
                "temperature": 0.2,
                "top_p": 0.8,
                "max_tokens": 400,
            },
            timeout=timeout
        )
        response.raise_for_status()
        result = response.json()
        response_text = result["choices"][0]["message"]["content"]
        log.info("Reponse Groq recue (%d caracteres)", len(response_text))
        return response_text
    except requests.exceptions.Timeout:
        log.error("TIMEOUT Groq apres %ds", timeout)
        return None
    except requests.exceptions.ConnectionError:
        log.error("Groq API non accessible sur %s", GROQ_API_URL)
        return None
    except Exception as e:
        log.error("Erreur Groq: %s", str(e))
        return None
        response = requests.post(
            GROQ_API_URL,
            json={
                "model": model,
                "prompt": prompt,
                "stream": False,
                "options": {
                    "temperature": 0.2,
                    "top_p": 0.8,
                    "num_ctx": 512,
                    "num_predict": 400,
                }
            },
            timeout=timeout
        )
        response.raise_for_status()
        result = response.json()
        response_text = result.get("response", "")
        log.info("✅ Reponse Groq recue (%d caracteres)", len(response_text))
        return response_text
    except requests.exceptions.Timeout:
        log.error(" TIMEOUT Groq apres %ds", timeout)
        return None
    except requests.exceptions.ConnectionError:
        log.error(" Groq non accessible sur %s", GROQ_API_URL)
        return None
    except Exception as e:
        log.error(" Erreur Groq: %s", str(e))
        return None


def _build_report_from_data(routes_data, total_tours, total_stops, total_km, total_h, avg_stops, avg_km, avg_h, eff_ratio):
    """
    Genere un rapport analytique complet en Python pur, sans IA.
    Utilise comme fallback quand Groq timeout ou est indisponible.
    """
    lines_out = []

    # 1. Synthese
    lines_out.append("## 📊 SYNTHESE GLOBALE")
    if avg_h < 4:
        charge_eval = "legeres (< 4h/tournee)"
    elif avg_h < 7:
        charge_eval = "bien dimensionnees (4-7h/tournee)"
    else:
        charge_eval = "chargees (> 7h/tournee)"
    lines_out.append(
        f"Le plan comprend {total_tours} tournees pour {total_stops} visites, "
        f"{total_km:.0f} km au total. Les tournees sont {charge_eval}. "
        f"Duree moyenne : {avg_h:.1f}h | Distance moyenne : {avg_km:.0f} km. "
        f"Vitesse commerciale estimee : {eff_ratio:.0f} km/h."
    )
    lines_out.append("")

    # 2. Equilibrage
    lines_out.append("##  EQUILIBRAGE DES CHARGES")
    overloaded = [r for r in routes_data if r['stops'] > avg_stops * 1.5]
    underloaded = [r for r in routes_data if r['stops'] < avg_stops * 0.5 and r['stops'] > 0]
    if overloaded:
        for r in overloaded:
            lines_out.append(f"  T{r['num']} ({r['city']}) : {r['stops']} arrets - SURCHARGEE (moy. {avg_stops:.0f}). Envisager de diviser.")
    if underloaded:
        for r in underloaded:
            lines_out.append(f"  T{r['num']} ({r['city']}) : {r['stops']} arrets - SOUS-CHARGEE (moy. {avg_stops:.0f}). Envisager de fusionner avec une tournee proche.")
    if not overloaded and not underloaded:
        lines_out.append(f"✅ Equilibrage satisfaisant - toutes les tournees sont proches de la moyenne ({avg_stops:.0f} arrets).")
    lines_out.append("")

    # 3. Nuitees
    lines_out.append("##  LOGISTIQUE & NUITEES")
    long_tours = [r for r in routes_data if r['time_hours'] > 8]
    if long_tours:
        for r in long_tours:
            lines_out.append(f"🌙 T{r['num']} ({r['city']}) : {r['time_hours']:.1f}h - Nuitee probable. Prevoir hebergement a {r['city']}.")
    else:
        max_h = max(r['time_hours'] for r in routes_data)
        lines_out.append(f"✅ Aucune nuitee necessaire - duree maximum : {max_h:.1f}h.")
    lines_out.append("")

    # 4. Recommandations
    lines_out.append("## 💡 RECOMMANDATIONS")
    reco_idx = 1

    if overloaded:
        names = ", ".join(f"T{r['num']}" for r in overloaded)
        lines_out.append(f"R{reco_idx}. Diviser les tournees surchargees : {names}")
        reco_idx += 1
    if underloaded:
        names = ", ".join(f"T{r['num']}" for r in underloaded)
        lines_out.append(f"R{reco_idx}. Fusionner les tournees sous-chargees : {names}")
        reco_idx += 1

    # Tournees avec fort temps de visite
    heavy_visit = [r for r in routes_data if r.get('visit_hours', 0) > r.get('travel_hours', 0) * 1.5]
    if heavy_visit:
        names = ", ".join(f"T{r['num']}" for r in heavy_visit[:3])
        lines_out.append(f"R{reco_idx}. Sites a temps de visite eleve ({names}) : les placer en debut de tournee pour optimiser la fatigue.")
        reco_idx += 1

    # Tournees longues en distance
    long_km = [r for r in routes_data if r['distance_km'] > avg_km * 1.5]
    if long_km:
        names = ", ".join(f"T{r['num']} ({r['distance_km']:.0f}km)" for r in long_km)
        lines_out.append(f"R{reco_idx}. Reduire la distance des tournees eloignees : {names} - verifier si des sites peuvent etre reaffectes a des tournees voisines.")
        reco_idx += 1

    if reco_idx == 1:
        lines_out.append("✅ Le plan est globalement bien optimise. Aucune recommandation critique.")
    lines_out.append("")

    # 5. Detail
    lines_out.append("## 🚨 TOURNEES CRITIQUES")
    critical = [r for r in routes_data if r['stops'] < 2 or r['time_hours'] > 8 or r['distance_km'] > 400 or r['stops'] > avg_stops * 1.8]
    if critical:
        for r in critical:
            issues = []
            if r['stops'] < 2: issues.append("trop peu d'arrets  fusionner")
            if r['time_hours'] > 8: issues.append(f"duree {r['time_hours']:.1f}h  prevoir nuitee ou diviser")
            if r['distance_km'] > 400: issues.append(f"distance {r['distance_km']:.0f}km  reaffecter sites lointains")
            if r['stops'] > avg_stops * 1.8: issues.append(f"{r['stops']} arrets  diviser en 2 tournees")
            lines_out.append(f"⚡ T{r['num']} ({r['city']}) : {' | '.join(issues)}")
    else:
        lines_out.append("✅ Aucune tournee critique detectee.")

    return "\n".join(lines_out)


def analyze_routes_with_ai(routes, sites, time_m, dist_m, depot_idx):
    """
    Analyse les tournees.
    Strategie : 1 seul appel Groq ultra-court, timeout 300s.
    Fallback Python pur si Groq timeout ou indisponible.
    """
    max_routes_to_analyze = min(len(routes), 20)
    routes_to_analyze = routes[:max_routes_to_analyze]
    log.info("🧠 Analyse : %d tournees", len(routes_to_analyze))

    # ── Calcul des metriques ───────────────────────────────────
    routes_data = []
    for i, route_data in enumerate(routes_to_analyze, 1):
        route = route_data["route"]
        stops = [sites[idx] for idx in route]
        total_distance = 0
        total_travel = 0
        for j in range(len(route)):
            curr = route[j]
            prev = depot_idx if j == 0 else route[j-1]
            total_distance += dist_m[prev][curr]
            total_travel += time_m[prev][curr]
        if route:
            last = route[-1]
            total_distance += dist_m[last][depot_idx]
            total_travel += time_m[last][depot_idx]
        main_city = stops[0].get("ville", "?") if stops else "?"
        visit_time = 0
        for s in stops:
            visit_time += _visit_minutes_from_row(s, default=30) * 60
        routes_data.append({
            "num": i,
            "stops": len(route),
            "distance_km": round(total_distance / 1000, 1),
            "time_hours": round((total_travel + visit_time) / 3600, 1),
            "travel_hours": round(total_travel / 3600, 1),
            "visit_hours": round(visit_time / 3600, 1),
            "city": main_city,
            "site_names": [s.get("site", "?") for s in stops],
        })

    total_tours  = len(routes_data)
    total_stops  = sum(r['stops'] for r in routes_data)
    total_km     = sum(r['distance_km'] for r in routes_data)
    total_h      = sum(r['time_hours'] for r in routes_data)
    avg_stops    = total_stops / max(1, total_tours)
    avg_km       = total_km / max(1, total_tours)
    avg_h        = total_h / max(1, total_tours)
    eff_ratio    = total_km / max(1, total_h)

    metrics = {
        "total_tours": total_tours, "total_stops": total_stops,
        "total_km": round(total_km, 1), "total_h": round(total_h, 1),
        "avg_stops": round(avg_stops, 1), "avg_km": round(avg_km, 1),
        "avg_h": round(avg_h, 1), "efficiency_kmh": round(eff_ratio, 1),
    }

    # ── Resume compact pour le prompt ─────────────────────────
    # On encode chaque tournee en une ligne tres courte pour tenir dans num_ctx=512
    short_summary = " | ".join(
        f"T{r['num']}:{r['city'][:8]},{r['stops']}arr,{r['distance_km']}km,{r['time_hours']}h"
        for r in routes_data
    )

    prompt = (
        f"Expert logistique. {total_tours} tournees, {total_stops} visites, "
        f"{total_km:.0f}km, {total_h:.0f}h. Moy:{avg_stops:.0f}arr/{avg_km:.0f}km/{avg_h:.1f}h. "
        f"Donnees: {short_summary}. "
        f"En 5-8 phrases concises: 1)synthese globale 2)tournees surchargees ou vides 3)nuitees si>8h "
        f"4)recommandations cles 5)tournees critiques. Pas de markdown, texte direct."
    )

    log.info("📤 Un seul appel Groq (prompt %d chars)", len(prompt))
    ai_text = ask_groq(prompt, timeout=60)

    # ── Fallback Python si Groq ne repond pas ────────────────
    if not ai_text:
        log.warning("  Groq indisponible - generation du rapport en mode Python pur")
        fallback_report = _build_report_from_data(
            routes_data, total_tours, total_stops, total_km, total_h,
            avg_stops, avg_km, avg_h, eff_ratio
        )
        return {
            "success": True,
            "raw_response": fallback_report,
            "analysis": {
                "global_assessment": fallback_report,
                "equilibrage": "",
                "nuites": "",
                "optimization": "",
                "details": "",
            },
            "metrics": metrics,
            "source": "python_fallback",
        }

    # ── Rapport avec reponse Groq ────────────────────────────
    # Egalement enrichi des donnees calculees en Python pour les metriques
    python_extras = _build_report_from_data(
        routes_data, total_tours, total_stops, total_km, total_h,
        avg_stops, avg_km, avg_h, eff_ratio
    )

    final_report = (
        "#  RAPPORT D'ANALYSE LOGISTIQUE\n\n"
        "## 🤖 ANALYSE IA\n" + ai_text +
        "\n\n---\n\n" + python_extras
    )

    log.info("✅ Rapport genere (IA + donnees calculees)")
    return {
        "success": True,
        "raw_response": final_report,
        "analysis": {
            "global_assessment": ai_text,
            "equilibrage": python_extras,
            "nuites": "",
            "optimization": "",
            "details": "",
        },
        "metrics": metrics,
        "source": "groq",
    }

def build_matrices(sites, profil):
    n = len(sites)
    BIG = 10_000_000
    time_m = [[BIG]*n for _ in range(n)]
    dist_m = [[BIG]*n for _ in range(n)]
    for i in range(n):
        time_m[i][i] = 0
        dist_m[i][i] = 0

    latlons = [(s["_lat"], s["_lon"]) for s in sites]

    if n <= CHUNK_SIZE:
        _fill_block(time_m, dist_m, list(range(n)), list(range(n)), latlons, profil)
    else:
        blocks = [list(range(i, min(i+CHUNK_SIZE, n))) for i in range(0, n, CHUNK_SIZE)]
        for bi in blocks:
            for bj in blocks:
                _fill_block(time_m, dist_m, bi, bj, latlons, profil)

    return time_m, dist_m

def _fill_block(time_m, dist_m, src_global, dst_global, all_latlons, profil):
    all_idx = list(dict.fromkeys(src_global + dst_global))
    local   = {g: li for li, g in enumerate(all_idx)}
    coords  = ";".join(f"{all_latlons[g][1]},{all_latlons[g][0]}" for g in all_idx)

    src_local = ";".join(str(local[g]) for g in src_global)
    dst_local = ";".join(str(local[g]) for g in dst_global)

    try:
        url = f"{OSRM_TABLE}/{profil}/{coords}"
        resp = SESSION.get(url, params={
            "sources": src_local,
            "destinations": dst_local,
            "annotations": "duration,distance",
        }, timeout=120)
        resp.raise_for_status()
        data = resp.json()

        if data.get("code") != "Ok":
            log.error("OSRM bloc erreur : %s", data.get("code"))
            return

        durs = data.get("durations") or []
        dsts = data.get("distances") or []

        for si, gi in enumerate(src_global):
            for di, gj in enumerate(dst_global):
                if gi == gj:
                    continue
                if durs and si < len(durs) and durs[si] and di < len(durs[si]):
                    v = durs[si][di]
                    if v is not None:
                        time_m[gi][gj] = int(v)
                if dsts and si < len(dsts) and dsts[si] and di < len(dsts[si]):
                    v = dsts[si][di]
                    if v is not None:
                        dist_m[gi][gj] = float(v)

    except Exception as e:
        log.error("OSRM call failed: %s", e)

# ─────────────────────────────────────────────────────────────────
# VRP OR-TOOLS avec prise en compte du temps par site
# ─────────────────────────────────────────────────────────────────
def run_vrp_robust(time_m, depot_idx, max_hours, solver_time_s, sites, max_s=14400):
    """
    Optimisation VRP basee sur le TEMPS TOTAL (trajet + temps sur site)
    Plutot que sur le nombre de sites.
    
    Args:
        time_m: Matrice des temps de trajet
        depot_idx: Index du depot
        max_hours: Duree maximale par tournee en heures
        solver_time_s: Temps max de recherche en secondes
        sites: Liste des sites avec leur temps de visite
        max_s: Temps maximum en secondes (fallback)
    """
    n = len(time_m)
    if n <= 1:
        return []
    
    # Convertir max_hours en secondes
    max_time_seconds = int(max_hours * 3600)
    log.info(f"    Contrainte temps par tournee: {max_hours}h ({max_time_seconds}s)")
    
    # Extraire les temps de visite des sites (en minutes)
    visit_times = []
    for i, site in enumerate(sites):
        if i == depot_idx:
            visit_times.append(0)  # Pas de temps de visite au depot
        else:
            visit_times.append(_visit_minutes_from_row(site, default=30) * 60)
    
    log.info(f"    Temps de visite moyens: {sum(visit_times)/len(visit_times)/60:.1f} min/site")
    
    # Creer une matrice de temps incluant les temps de visite
    # Pour chaque arc (i->j), le temps total = trajet(i->j) + temps_visite(i)
    # Sauf pour le depot qui a un temps de visite = 0
    augmented_time_matrix = []
    for i in range(n):
        row = []
        for j in range(n):
            if i == j:
                row.append(0)
            else:
                # Temps de trajet + temps de visite sur le site i
                row.append(time_m[i][j] + visit_times[i])
        augmented_time_matrix.append(row)
    
    # Estimer le nombre de vehicules necessaires
    # Somme des temps de visite / temps max par vehicule
    total_visit_time = sum(visit_times)
    estimated_vehicles = max(1, min(n-1, int(math.ceil(total_visit_time / max_time_seconds * 1.2))))
    log.info(f"  🚚  Estimation vehicules: {estimated_vehicles}")
    
    data = {
        'time_matrix': augmented_time_matrix,
        'num_vehicles': estimated_vehicles,
        'depot': depot_idx,
        'max_time_per_vehicle': max_time_seconds
    }

    manager = pywrapcp.RoutingIndexManager(len(data['time_matrix']),
                                           data['num_vehicles'], data['depot'])
    routing = pywrapcp.RoutingModel(manager)

    # Dimension temps total (trajet + visite)
    def time_callback(from_index, to_index):
        from_node = manager.IndexToNode(from_index)
        to_node = manager.IndexToNode(to_index)
        return data['time_matrix'][from_node][to_node]

    transit_callback_index = routing.RegisterTransitCallback(time_callback)
    routing.AddDimension(
        transit_callback_index,
        0,  # temps d'attente maximum (0 = pas d'attente)
        max_time_seconds,  # temps maximum par vehicule
        True,  # commencer a 0 (au depot)
        'Time'
    )
    
    # Fixer le cout a minimiser sur la dimension Temps
    time_dimension = routing.GetDimensionOrDie('Time')
    routing.SetArcCostEvaluatorOfAllVehicles(transit_callback_index)

    # Recherche
    search_parameters = pywrapcp.DefaultRoutingSearchParameters()
    search_parameters.first_solution_strategy = (
        routing_enums_pb2.FirstSolutionStrategy.PATH_CHEAPEST_ARC)
    search_parameters.time_limit.seconds = solver_time_s
    # Utiliser un guided local search pour de meilleurs resultats
    search_parameters.local_search_metaheuristic = (
        routing_enums_pb2.LocalSearchMetaheuristic.GUIDED_LOCAL_SEARCH)
    search_parameters.solution_limit = 100

    solution = routing.SolveWithParameters(search_parameters)

    if solution:
        routes = []
        for vehicle_id in range(data['num_vehicles']):
            index = routing.Start(vehicle_id)
            route = []
            while not routing.IsEnd(index):
                node_index = manager.IndexToNode(index)
                if node_index != depot_idx:
                    route.append(node_index)
                index = solution.Value(routing.NextVar(index))
            if len(route) > 0:
                routes.append(route)
        
        log.info(f"  ✅ Solution trouvee: {len(routes)} tournees")
        return routes
    else:
        log.warning("  Pas de solution avec contrainte de temps, tentative sans contrainte")
        # Fallback: sans contrainte de temps stricte
        return _run_vrp_no_constraint(time_m, depot_idx, max_hours, sites)

def _run_vrp_no_constraint(time_m, depot_idx, max_hours, sites):
    """
    Fallback: VRP simple sans contrainte de temps stricte
    Utilise quand la solution avec contrainte echoue
    """
    n = len(time_m)
    if n <= 1:
        return []
    
    # Temps de visite en secondes
    visit_times = []
    for i, site in enumerate(sites):
        if i == depot_idx:
            visit_times.append(0)
        else:
            visit_times.append(_visit_minutes_from_row(site, default=30) * 60)
    
    # Estimer nombre de vehicules
    total_visit_time = sum(visit_times)
    max_time_seconds = int(max_hours * 3600)
    estimated_vehicles = max(1, min(n-1, int(math.ceil(total_visit_time / max_time_seconds * 1.3))))
    
    data = {
        'time_matrix': time_m,
        'num_vehicles': estimated_vehicles,
        'depot': depot_idx
    }

    manager = pywrapcp.RoutingIndexManager(len(data['time_matrix']),
                                           data['num_vehicles'], data['depot'])
    routing = pywrapcp.RoutingModel(manager)

    def time_callback(from_index, to_index):
        from_node = manager.IndexToNode(from_index)
        to_node = manager.IndexToNode(to_index)
        return data['time_matrix'][from_node][to_node]

    transit_callback_index = routing.RegisterTransitCallback(time_callback)
    routing.SetArcCostEvaluatorOfAllVehicles(transit_callback_index)

    search_parameters = pywrapcp.DefaultRoutingSearchParameters()
    search_parameters.first_solution_strategy = (
        routing_enums_pb2.FirstSolutionStrategy.PATH_CHEAPEST_ARC)
    search_parameters.time_limit.seconds = 45

    solution = routing.SolveWithParameters(search_parameters)

    if not solution:
        return []

    routes = []
    for vehicle_id in range(data['num_vehicles']):
        index = routing.Start(vehicle_id)
        route = []
        while not routing.IsEnd(index):
            node_index = manager.IndexToNode(index)
            if node_index != depot_idx:
                route.append(node_index)
            index = solution.Value(routing.NextVar(index))
        if len(route) > 0:
            routes.append(route)
    
    return routes



def run_vrp_with_date_groups(time_m, dist_m, sites, depot_idx, date_groups, group_size, solver_time_s):
    """
    Optimisation VRP avec contraintes de fenetres temporelles
    Chaque groupe de dates est optimise separement
    """
    all_routes = []
    
    for group_idx, group in enumerate(date_groups):
        # Indices des sites dans ce groupe (indices globaux)
        group_site_indices = []
        
        # CORRECTION ICI : Parcourir les items du groupe
        for item in group["sites"]:
            # Gerer le mix entre objets wrappes et lignes brutes
            if isinstance(item, dict) and "site" in item:
                target = item["site"]
            else:
                target = item
            
            site_name = target.get("site", "")
            
            # Trouver l'index global dans la liste complete des sites
            for global_idx, s in enumerate(sites):
                if s.get("site", "") == site_name:
                    group_site_indices.append(global_idx)
                    break
        
        # Ajouter le depot si pas deja present
        if depot_idx not in group_site_indices:
            group_site_indices.append(depot_idx)
        
        group_n = len(group_site_indices)
        
        # Creer une matrice reduite pour ce groupe
        # Il faut mapper les indices globaux vers des indices locaux (0..N-1)
        # pour la creation de la sous-matrice
        # On utilise simplement l'ordre dans group_site_indices
        local_indices_map = {global_idx: local_idx for local_idx, global_idx in enumerate(group_site_indices)}
        
        group_time_m = [[time_m[i][j] for j in group_site_indices] for i in group_site_indices]
        
        log.info(f"📅 Optimisation groupe {group_idx + 1}/{len(date_groups)} : {group_n} sites (dont depot)")
        
        # Creer la liste de sites pour ce groupe (dans l'ordre group_site_indices)
        group_sites_list = [sites[i] for i in group_site_indices]
        
        # Optimiser ce groupe
        # Dans la matrice reduite, le depot est toujours a la position group_site_indices.index(depot_idx)
        local_depot_idx = group_site_indices.index(depot_idx)
        
        local_routes = run_vrp_robust(group_time_m, local_depot_idx, group_size, solver_time_s, group_sites_list)
        
        # Convertir les indices locaux en indices globaux
        for route in local_routes:
            global_route = [group_site_indices[idx] for idx in route if idx != local_depot_idx]
            all_routes.append({
                "route": global_route,
                "group_idx": group_idx,
                "date_min": group.get("date_min"),
                "date_max": group.get("date_max")
            })
    
    log.info(f"📅 {len(all_routes)} tournees totales generees")
    return all_routes





def split_long_tours_by_time(routes, time_m, depot_idx, max_s, split_size, sites, split_mode="manual", dist_m=None, use_ai=False):
    """
    Split tours with 2 modes:
    - manual: fixed size (split_size), with minimum 2 sites per tour
    - auto: variable size, tours are filled until max time is reached.
            No artificial cap on sites/tour  only time + date constraints apply.
    """
    split_mode = str(split_mode or "manual").strip().lower()
    if split_mode not in ("manual", "auto"):
        split_mode = "manual"

    use_ai = bool(use_ai)
    ai_merge_budget = 3
    min_sites = 2
    fixed_size = max(min_sites, int(split_size or min_sites))
    # En mode auto, pas de plafond artificiel : les tournes se remplissent
    # selon le temps disponible (max_s). Le seul plafond est le nombre total de sites.
    auto_max_sites = max(len(sites), 50) if split_mode == "auto" else 8

    def _to_datetime_loose(value):
        d = parse_date(value)
        if d:
            return d
        txt = str(value or "").strip()
        if not txt:
            return None
        try:
            dt = pd.to_datetime(txt, dayfirst=True, errors="coerce")
            if pd.notna(dt):
                return dt.to_pydatetime()
        except Exception:
            pass
        return None

    def _site_window(site_idx):
        s = sites[site_idx]
        d_min = _to_datetime_loose(s.get("date_min", "")) or _to_datetime_loose(s.get("_date_min", ""))
        d_max = _to_datetime_loose(s.get("date_max", "")) or _to_datetime_loose(s.get("_date_max", ""))
        return d_min, d_max

    def _merge_windows(curr_min, curr_max, new_min, new_max):
        merged_min = curr_min
        merged_max = curr_max
        if new_min and (merged_min is None or new_min > merged_min):
            merged_min = new_min
        if new_max and (merged_max is None or new_max < merged_max):
            merged_max = new_max
        is_ok = not (merged_min and merged_max and merged_min > merged_max)
        return is_ok, merged_min, merged_max

    def _route_travel_seconds(route):
        if not route:
            return 0
        seq = [depot_idx] + route + [depot_idx]
        return sum(time_m[seq[i]][seq[i + 1]] for i in range(len(seq) - 1))

    def _route_total_seconds(route):
        if not route:
            return 0
        visits = sum(_visit_minutes_from_row(sites[idx], default=30) * 60 for idx in route)
        return _route_travel_seconds(route) + visits

    def _route_metric(route):
        """Cot de minimisation: km si dist_m disponible, sinon temps de trajet."""
        if not route:
            return 0.0
        seq = [depot_idx] + route + [depot_idx]
        matrix = dist_m if dist_m is not None else time_m
        return float(sum(matrix[seq[i]][seq[i + 1]] for i in range(len(seq) - 1)))

    def _route_metric_human(route):
        val = _route_metric(route)
        return (val / 1000.0) if dist_m is not None else (val / 3600.0)

    def _route_dates_compatible(route):
        win_min, win_max = None, None
        for idx in route:
            d_min, d_max = _site_window(idx)
            ok, win_min, win_max = _merge_windows(win_min, win_max, d_min, d_max)
            if not ok:
                return False
        return True

    def _valid_chunk(route, min_sz, max_sz, check_time=True, check_dates=True):
        if not route:
            return False
        if len(route) < min_sz or len(route) > max_sz:
            return False
        if check_time and len(route) > 1 and _route_total_seconds(route) > max_s:
            return False
        if check_dates and not _route_dates_compatible(route):
            return False
        return True

    def _has_small_chunk(chunks, min_sz):
        return any(len(c) < min_sz for c in chunks if c)

    def _chunk_penalty(route, max_sz, strict=True):
        if not route:
            return float("inf")
        p = 0.0
        if len(route) > max_sz:
            p += (len(route) - max_sz) * 1e9
        if strict:
            if len(route) > 1 and _route_total_seconds(route) > max_s:
                p += (_route_total_seconds(route) - max_s) * 3e4
            if not _route_dates_compatible(route):
                p += 2e9
        return p

    def _merge_small_chunks(chunks, min_sz, max_sz, strict=True):
        chunks = [c for c in chunks if c]
        total_sites = sum(len(c) for c in chunks)
        if total_sites < min_sz:
            return chunks

        guard = 0
        while _has_small_chunk(chunks, min_sz) and len(chunks) > 1 and guard < 500:
            guard += 1
            small_idx = next((i for i, c in enumerate(chunks) if len(c) < min_sz), None)
            if small_idx is None:
                break

            small = chunks[small_idx]
            best_move = None  # (score, donor_idx, donor_new, small_new)

            # 1) Essayer d'emprunter un site d'un voisin > min_sz pour atteindre min_sz
            for j, donor in enumerate(chunks):
                if j == small_idx or len(donor) <= min_sz:
                    continue

                donor_options = []
                if donor:
                    donor_options.append((donor[:-1], donor[-1]))  # prendre la fin
                    donor_options.append((donor[1:], donor[0]))    # prendre le dbut

                for donor_new, moved in donor_options:
                    for small_new in (small + [moved], [moved] + small):
                        if len(donor_new) < min_sz or len(small_new) > max_sz:
                            continue

                        p = _chunk_penalty(donor_new, max_sz, strict) + _chunk_penalty(small_new, max_sz, strict)
                        if p >= 1e11:
                            continue
                        delta = (_route_metric(donor_new) + _route_metric(small_new)) - (_route_metric(donor) + _route_metric(small))
                        score = p + delta
                        if best_move is None or score < best_move[0]:
                            best_move = (score, j, donor_new, small_new)

            if best_move is not None:
                _, donor_idx, donor_new, small_new = best_move
                chunks[donor_idx] = donor_new
                chunks[small_idx] = small_new
                chunks = [c for c in chunks if c]
                continue

            # 2) Sinon fusionner le petit morceau avec la meilleure tourne voisine
            best_merge = None  # (score, base_idx, merged_chunk)
            for j, base in enumerate(chunks):
                if j == small_idx:
                    continue
                for merged in (base + small, small + base):
                    if len(merged) > max_sz:
                        continue
                    p = _chunk_penalty(merged, max_sz, strict)
                    if p >= 1e11:
                        continue
                    delta = _route_metric(merged) - _route_metric(base) - _route_metric(small)
                    score = p + delta
                    if best_merge is None or score < best_merge[0]:
                        best_merge = (score, j, merged)

            if best_merge is None:
                break

            _, base_idx, merged_chunk = best_merge
            if small_idx < base_idx:
                chunks.pop(small_idx)
                base_idx -= 1
                chunks[base_idx] = merged_chunk
            else:
                chunks[base_idx] = merged_chunk
                chunks.pop(small_idx)

        if _has_small_chunk(chunks, min_sz):
            log.warning(f"       Morceau < {min_sz} sites conserv (mode {'strict' if strict else 'relax'})")
        return chunks

    def _manual_split_route(route):
        if not route:
            return []
        chunks = [route[i:i + fixed_size] for i in range(0, len(route), fixed_size)]

        # viter un dernier morceau  1 site
        if len(chunks) >= 2 and len(chunks[-1]) < min_sites:
            if len(chunks[-2]) > min_sites:
                chunks[-1].insert(0, chunks[-2].pop())
            else:
                chunks[-2].extend(chunks[-1])
                chunks.pop()
                log.info("      Ajustement min 2 sites (fusion du dernier morceau)")
        return chunks

    def _dp_partition_auto(route):
        """Partition dynamique minimisant le cot total (km/temps) sous contraintes."""
        n = len(route)
        inf = float("inf")
        dp = [inf] * (n + 1)
        nxt = [-1] * (n + 1)
        dp[n] = 0.0

        for i in range(n - 1, -1, -1):
            max_j = min(n, i + auto_max_sites)
            for j in range(i + min_sites, max_j + 1):
                seg = route[i:j]
                if not _valid_chunk(seg, min_sites, auto_max_sites, check_time=True, check_dates=True):
                    continue

                # Minimiser le cot de trajet pur (km ou temps) + pnalit par tourne
                # pour favoriser des tournes plus remplies (moins de tournes au total)
                tour_overhead = 50_000.0 if dist_m is not None else 3_600.0
                seg_cost = _route_metric(seg) + tour_overhead
                cand = seg_cost + dp[j]
                if cand < dp[i]:
                    dp[i] = cand
                    nxt[i] = j

        if nxt[0] == -1:
            return None

        parts = []
        i = 0
        while i < n:
            j = nxt[i]
            if j <= i:
                return None
            parts.append(route[i:j])
            i = j
        return parts

    def _greedy_partition_auto(route):
        """Fallback robuste si la DP ne trouve pas de partition stricte."""
        chunks = []
        current = []
        cur_min, cur_max = None, None

        for node_idx in route:
            n_min, n_max = _site_window(node_idx)
            if not current:
                current = [node_idx]
                cur_min, cur_max = n_min, n_max
                continue

            date_ok, cand_min, cand_max = _merge_windows(cur_min, cur_max, n_min, n_max)
            cand = current + [node_idx]
            should_split = (
                len(cand) > auto_max_sites
                or (len(cand) > 1 and _route_total_seconds(cand) > max_s)
                or (not date_ok)
            )

            if should_split:
                chunks.append(current)
                current = [node_idx]
                cur_min, cur_max = n_min, n_max
            else:
                current = cand
                cur_min, cur_max = cand_min, cand_max

        if current:
            chunks.append(current)
        return chunks

    def _order_nodes_greedy(nodes):
        """Greedy ordering from depot to estimate merged route quality."""
        rem = list(dict.fromkeys(nodes))
        if len(rem) <= 2:
            return rem
        matrix = dist_m if dist_m is not None else time_m
        cur = depot_idx
        ordered = []
        while rem:
            nxt = min(rem, key=lambda k: matrix[cur][k])
            ordered.append(nxt)
            rem.remove(nxt)
            cur = nxt
        return ordered

    def _choose_merge_with_ai(candidates):
        """
        candidates: list[dict] sorted by descending saving.
        Returns selected candidate (or best deterministic fallback).
        """
        if not candidates:
            return None
        nonlocal ai_merge_budget
        if not use_ai or not GROQ_API_KEY or ai_merge_budget <= 0:
            return candidates[0]

        top = candidates[: min(4, len(candidates))]
        unit = "km" if dist_m is not None else "h_trajet"
        lines = []
        for idx, c in enumerate(top, 1):
            lines.append(
                f"{idx}) merge T{c['i']+1}+T{c['j']+1} -> {len(c['merged'])} sites, "
                f"gain={c['saving_human']:.2f} {unit}, total={c['merged_total_h']:.2f}h"
            )

        prompt = (
            "Tu es un optimiseur VRP. Choisis UNE seule option qui minimise le kilometrage global "
            "sans depasser les contraintes de temps/date deja verifiees.\n"
            "Reponds uniquement par le numero (1..4).\n\n"
            + "\n".join(lines)
        )
        try:
            ai_merge_budget -= 1
            ans = ask_groq(prompt, timeout=15)
            if ans:
                m = re.search(r"\b([1-4])\b", str(ans))
                if m:
                    k = int(m.group(1)) - 1
                    if 0 <= k < len(top):
                        return top[k]
        except Exception:
            pass
        return candidates[0]

    def _coalesce_routes_for_km(chunks):
        """
        Merge compatible tours to reduce total metric (km/time).
        Only for auto mode.
        """
        chunks = [c for c in chunks if c]
        if len(chunks) <= 1:
            return chunks

        while True:
            candidates = []
            for i in range(len(chunks)):
                for j in range(i + 1, len(chunks)):
                    a = chunks[i]
                    b = chunks[j]
                    if len(a) + len(b) > auto_max_sites:
                        continue

                    merged = _order_nodes_greedy(a + b)
                    if not _valid_chunk(merged, min_sites, auto_max_sites, check_time=True, check_dates=True):
                        continue

                    before = _route_metric(a) + _route_metric(b)
                    after = _route_metric(merged)
                    saving = before - after
                    if saving <= 0:
                        continue

                    candidates.append({
                        "i": i,
                        "j": j,
                        "merged": merged,
                        "saving": saving,
                        "saving_human": (_route_metric_human(a) + _route_metric_human(b) - _route_metric_human(merged)),
                        "merged_total_h": _route_total_seconds(merged) / 3600.0,
                    })

            if not candidates:
                break

            candidates.sort(key=lambda c: (c["saving"], len(c["merged"])), reverse=True)
            chosen = _choose_merge_with_ai(candidates)
            if not chosen:
                break

            i, j = chosen["i"], chosen["j"]
            a_len, b_len = len(chunks[i]), len(chunks[j])
            chunks[i] = chosen["merged"]
            del chunks[j]
            unit = "km" if dist_m is not None else "h trajet"
            log.info(
                f"      Fusion km: {a_len}+{b_len}->{len(chosen['merged'])} sites, "
                f"gain {chosen['saving_human']:.2f} {unit}"
            )

        return chunks

    if split_mode == "manual":
        final = []
        for r in routes:
            if not r:
                continue
            total_time = _route_total_seconds(r)
            if len(r) > fixed_size:
                log.info(f"    Dcoupage manuel fixe: {len(r)} sites, {total_time/3600:.1f}h (taille {fixed_size})")
            chunks = _manual_split_route(r)
            for ch in chunks:
                ch_time = _route_total_seconds(ch)
                if ch_time > max_s and len(ch) > 1:
                    log.warning(f"       Morceau fixe long: {len(ch)} sites, {ch_time/3600:.1f}h > {max_s/3600:.1f}h")
                log.info(f"      Nouveau morceau fixe: {len(ch)} sites, {ch_time/3600:.1f}h")
            final.extend(chunks)
        final = _merge_small_chunks(final, min_sites, max(fixed_size, min_sites), strict=True)
        if _has_small_chunk(final, min_sites):
            final = _merge_small_chunks(final, min_sites, max(fixed_size, min_sites), strict=False)
        return final

    # Mode auto: dcoupage variable  chaque tourne est remplie au maximum du temps disponible
    final = []
    for r in routes:
        if not r:
            continue

        total_time = _route_total_seconds(r)
        needs_split = (
            len(r) > auto_max_sites
            or total_time > max_s
            or (not _route_dates_compatible(r))
        )

        if not needs_split and len(r) >= min_sites:
            final.append(r)
            continue

        log.info(
            f"    Dcoupage auto intelligent: {len(r)} sites, {total_time/3600:.1f}h "
            f"(max {max_s/3600:.1f}h/tourne, min {min_sites} sites)"
        )

        chunks = _dp_partition_auto(r)
        if not chunks:
            chunks = _greedy_partition_auto(r)

        chunks = _merge_small_chunks(chunks, min_sites, auto_max_sites, strict=True)
        if _has_small_chunk(chunks, min_sites):
            chunks = _merge_small_chunks(chunks, min_sites, auto_max_sites, strict=False)

        # Reduce number of tours when possible by merging routes that lower total km/time
        chunks = _coalesce_routes_for_km(chunks)
        chunks = _merge_small_chunks(chunks, min_sites, auto_max_sites, strict=True)
        if _has_small_chunk(chunks, min_sites):
            chunks = _merge_small_chunks(chunks, min_sites, auto_max_sites, strict=False)

        for ch in chunks:
            ch_time = _route_total_seconds(ch)
            if ch_time > max_s and len(ch) > 1:
                log.warning(f"       Morceau auto long: {len(ch)} sites, {ch_time/3600:.1f}h > {max_s/3600:.1f}h")
            log.info(f"      Nouveau morceau: {len(ch)} sites, {ch_time/3600:.1f}h")
        final.extend(chunks)

    # Dernire passe globale pour liminer les morceaux  1 site quand possible
    final = _merge_small_chunks(final, min_sites, auto_max_sites, strict=True)
    if _has_small_chunk(final, min_sites):
        final = _merge_small_chunks(final, min_sites, auto_max_sites, strict=False)
    final = _coalesce_routes_for_km(final)
    final = _merge_small_chunks(final, min_sites, auto_max_sites, strict=True)
    if _has_small_chunk(final, min_sites):
        final = _merge_small_chunks(final, min_sites, auto_max_sites, strict=False)
    return final
# ─────────────────────────────────────────────────────────────────
# EXCEL
# ─────────────────────────────────────────────────────────────────
PALETTE = [
    "E3F2FD","E8F5E9","FFF3E0","F3E5F5","E0F7FA",
    "FCE4EC","F9FBE7","EDE7F6","FFFDE7","E0F2F1",
]

def build_excel(tours_data):
    """Genere le fichier Excel des tournees"""
    cols = ["tour_num","stop_order","tour_total_distance_km",
            "tour_total_travel_h","site","client","ville","address"]

    wb = Workbook()
    ws = wb.active
    ws.title = "tours"
    ws.append(cols)

    hfill = PatternFill("solid", fgColor="1F2937")
    hfont = Font(color="FFFFFF", bold=True)
    for ci, h in enumerate(cols, 1):
        c = ws.cell(row=1, column=ci)
        c.fill = hfill
        c.font = hfont
        c.alignment = Alignment(horizontal="center", vertical="center")

    row_idx = 2
    for td in tours_data:
        tnum = td["tour_num"]
        fill = PatternFill("solid", fgColor=PALETTE[(tnum-1) % len(PALETTE)])
        for order, stop in enumerate(td["stops"], 1):
            ws.append([tnum, order, td["total_km"], td["total_h"],
                       stop.get("site",""), stop.get("client",""),
                       stop.get("ville",""), stop.get("address","")])
            for ci in range(1, len(cols)+1):
                ws.cell(row=row_idx, column=ci).fill = fill
                ws.cell(row=row_idx, column=ci).alignment = Alignment(vertical="center")
            row_idx += 1

    for ci, h in enumerate(cols, 1):
        mx = len(h)
        for ri in range(2, ws.max_row+1):
            v = ws.cell(row=ri, column=ci).value
            if v is not None:
                mx = max(mx, len(str(v)))
        ws.column_dimensions[get_column_letter(ci)].width = min(mx+2, 80)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode()

# ─────────────────────────────────────────────────────────────────
# PIPELINE COMPLET
# ─────────────────────────────────────────────────────────────────
def parse_date(date_str):
    """
    Convertit une date string en datetime
    Accepte plusieurs formats : DD/MM/YYYY, YYYY-MM-DD, DD/MM/YY, etc.
    """
    if not date_str or str(date_str).strip() in ("", "0", "nan", "None"):
        return None
    
    date_str = str(date_str).strip()
    
    # Essayer differents formats
    formats = [
        "%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y", "%Y%m%d",
        "%d-%m-%Y", "%Y%m%d%H%M", "%d/%m/%Y %H:%M"
    ]
    
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    
    return None

def get_date_windows(sites):
    """
    Analyse les sites et identifie les intervalles de dates compatibles
    Retourne une liste de groupes avec leurs dates communes
    """
    date_windows = []
    
    # Filtrer les sites qui ont des dates valides
    dated_sites = []
    for site in sites:
        date_min = parse_date(site.get("date_min", ""))
        date_max = parse_date(site.get("date_max", ""))
        
        if date_min and date_max:
            dated_sites.append({
                "site": site,
                "date_min": date_min,
                "date_max": date_max,
                "_index": sites.index(site)
            })
    
    if not dated_sites:
        log.info("📅 Aucune date trouvee, creation d'un groupe unique")
        return [{"sites": sites, "date_min": None, "date_max": None}]
    
    log.info(f"📅 {len(dated_sites)} sites avec dates sur {len(sites)} sites totaux")
    
    # Trier par date_min
    dated_sites.sort(key=lambda x: x["date_min"] or datetime.min)
    
    # Creer des groupes de sites compatibles (intervles qui se chevauchent)
    groups = []
    current_group = {"sites": [], "date_min": None, "date_max": None}
    
    for site in dated_sites:
        site_min = site["date_min"]
        site_max = site["date_max"]
        
        if not current_group["sites"]:
            # Premier site du groupe
            current_group["sites"].append(site)
            current_group["date_min"] = site_min
            current_group["date_max"] = site_max
        else:
            # Verifier compatibilite avec le groupe actuel
            group_min = current_group["date_min"]
            group_max = current_group["date_max"]
            
            # Verifier si les intervalles se chevauchent
            intervals_overlap = (
                (site_min <= group_max + pd.Timedelta(days=1)) and 
                (site_max >= group_min - pd.Timedelta(days=1))
            )
            
            if intervals_overlap:
                # Ajouter au groupe existant et etendre l'intervalle
                current_group["sites"].append(site)
                if site_min < current_group["date_min"]:
                    current_group["date_min"] = site_min
                if site_max > current_group["date_max"]:
                    current_group["date_max"] = site_max
            else:
                # Sauvegarder le groupe actuel et en creer un nouveau
                if current_group["sites"]:
                    groups.append(current_group.copy())
                current_group = {
                    "sites": [site],
                    "date_min": site_min,
                    "date_max": site_max
                }
    
    # Ajouter le dernier groupe s'il a des sites
    if current_group["sites"]:
        groups.append(current_group)
    
    log.info(f"📅 {len(groups)} groupes de dates identifies")
    
    # Ajouter les sites sans dates au premier groupe
    sites_without_dates = [s for s in sites if "date_min" not in s and "date_max" not in s]
    
    if sites_without_dates and groups:
        log.info(f"📅 {len(sites_without_dates)} sites sans dates ajoutes au premier groupe")
        groups[0]["sites"].extend(sites_without_dates)
    elif not groups and sites_without_dates:
        # Si aucun groupe avec dates, creer un groupe unique
        groups = [{"sites": sites, "date_min": None, "date_max": None}]
    
    return groups

def smart_merge_routes(routes, sites, time_m, dist_m, depot_idx, min_sites, max_sites, max_time):
    """
    Fusionne les tournees trop courtes avec leurs voisines geographiques
    pour respecter le minimum de sites, sans depasser les contraintes.
    """
    merged = True
    # On travaille sur une copie des routes (liste de dicts)
    current_routes = [{"route": list(r["route"])} for r in routes]
    
    log.info(f"  🔄 Smart Merge : {len(current_routes)} tournees initiales (Min sites: {min_sites})")

    while merged:
        merged = False
        
        # Trouver les tournees trop courtes
        small_tour_indices = [i for i, r in enumerate(current_routes) if len(r["route"]) < min_sites]
        
        if not small_tour_indices:
            break
            
        # Tenter de fusionner chaque petite tournee
        for idx in small_tour_indices:
            if idx >= len(current_routes): continue # Deja supprimee ?
            if len(current_routes[idx]["route"]) >= min_sites: continue # Deja fusionnee dans ce tour de boucle
            
            route_a = current_routes[idx]["route"]
            if not route_a: continue # Route vide ?

            # Trouver la tournee la plus proche geographiquement (simple heuristique : fin de A proche du debut de B)
            best_neighbor_idx = -1
            min_dist_to_b = float('inf')
            
            # Dernier site de la tournee A
            last_node_a = route_a[-1]
            coord_a = (sites[last_node_a]["_lat"], sites[last_node_a]["_lon"])
            
            for j, route_dict_b in enumerate(current_routes):
                if idx == j: continue
                
                route_b = route_dict_b["route"]
                # Verifier capacite
                if len(route_a) + len(route_b) > max_sites:
                    continue
                    
                # Premier site de la tournee B
                first_node_b = route_b[0]
                coord_b = (sites[first_node_b]["_lat"], sites[first_node_b]["_lon"])
                
                # Calcul distance simple (Euclidienne) pour trouver le voisin le plus proche
                # C'est une heuristique rapide avant de verifier le temps reel
                dist = ((coord_a[0]-coord_b[0])**2 + (coord_a[1]-coord_b[1])**2)**0.5
                
                if dist < min_dist_to_b:
                    min_dist_to_b = dist
                    best_neighbor_idx = j
            
            # Si un voisin est trouve, verifier si la fusion est valide en temps
            if best_neighbor_idx != -1:
                route_b = current_routes[best_neighbor_idx]["route"]
                
                # Simuler le temps : A + (Distance A->B) + B
                # On calcule le temps de A complet, B complet, et le lien entre eux
                def calc_dur(route_nodes):
                    seq = [depot_idx] + route_nodes + [depot_idx]
                    return sum(time_m[seq[k]][seq[k+1]] for k in range(len(seq)-1))
                
                time_a = calc_dur(route_a)
                time_b = calc_dur(route_b)
                
                # Temps de transition entre fin de A et debut de B
                time_trans = time_m[last_node_a][route_b[0]]
                
                total_merged_time = time_a + time_trans + time_b
                
                if total_merged_time <= max_time:
                    # FUSION VALIDEE
                    # Nouvelle route : A -> B (ou B -> A, ici on concatene simplement)
                    # Note: L'ordre optimal serait A puis B (on finit A, on va au debut de B)
                    new_route = route_a + route_b
                    
                    # On remplace A par la fusion, on supprime B
                    current_routes[idx]["route"] = new_route
                    del current_routes[best_neighbor_idx]
                    
                    log.info(f"  ✅ Fusion : Tournee {idx+1} ({len(route_a)} sites) + Tournee {best_neighbor_idx+1} ({len(route_b)} sites)")
                    merged = True
                    break # On sort de la boucle pour recommencer l'analyse

    # Reconvertir en format attendu par la suite (avec metadonnees vides pour les nouvelles routes)
    final_routes = []
    for r_dict in current_routes:
        final_routes.append({
            "route": r_dict["route"],
            "group_idx": 0, # Plus de groupe de date
            "date_min": None,
            "date_max": None
        })
        
    log.info(f"   Smart Merge termine : {len(final_routes)} tournees finales.")
    return final_routes


def full_pipeline_enhanced(filepath, params, session_id, progress_id=""):
    depot_name   = params["depot_site"]
    max_hours    = float(params.get("max_tour_hours", 7.0))  # Nouveau parametre: duree max par tournee
    solver_time  = params["solver_time"]
    max_s        = int(max_hours * 3600)  # En secondes
    split_size   = params["split_size"]
    split_mode   = str(params.get("split_mode", "manual")).strip().lower()
    if split_mode not in ("manual", "auto"):
        split_mode = "manual"
    profil       = params["profil"]
    use_ai       = params.get("use_ai", True)

    def _prog(pct, msg):
        set_progress(progress_id, pct, msg)

    log.info(f"  Mode planification base sur le TEMPS (max {max_hours}h/tournee)")

    # 1. Lecture
    log.info("=== 1/7 Lecture Excel ===")
    _prog(5, "Lecture du fichier...")
    df = pd.read_excel(filepath, dtype=str, na_filter=False, keep_default_na=False)
    if df.empty:
        raise ValueError("Le fichier Excel est vide.")
    
    # Colonnes optionnelles avec valeurs par defaut
    for col in ["site", "address", "ville"]:
        if col not in df.columns:
            df[col] = ""
    
    # Colonnes de dates
    if "date_min" not in df.columns:
        df["date_min"] = ""
    if "date_max" not in df.columns:
        df["date_max"] = ""
    
    # Colonne temps par site - supporte "temps_site" (minutes) ET "Temps" (heures)
    def _heures_to_minutes(v):
        try:
            h = float(str(v).strip().replace(",", "."))
            return str(int(round(h * 60))) if h > 0 else "30"
        except (ValueError, TypeError):
            return "30"

    if "temps_site" not in df.columns:
        df["temps_site"] = ""

    if "Temps" in df.columns:
        # Priorit  la colonne mtier 'Temps' (en heures) si elle est renseigne.
        hours_mask = df["Temps"].astype(str).str.strip().ne("")
        if hours_mask.any():
            df.loc[hours_mask, "temps_site"] = df.loc[hours_mask, "Temps"].apply(_heures_to_minutes)
            log.info("    Colonne 'Temps' (heures) applique  %d lignes (prioritaire)", int(hours_mask.sum()))
    else:
        # Heuristique: si seules des petites valeurs existent dans temps_site, elles sont probablement en heures.
        ts_num = pd.to_numeric(df["temps_site"].astype(str).str.replace(",", ".", regex=False), errors="coerce")
        ts_valid = ts_num.dropna()
        if len(ts_valid) >= 3 and ts_valid.max() <= 12 and ts_valid.median() <= 8:
            df["temps_site"] = ts_num.apply(
                lambda x: str(int(round(x * 60))) if pd.notna(x) and x > 0 else ""
            )
            log.warning("    'Temps' absent: 'temps_site' ressemble  des heures, conversion auto heuresminutes applique")

    empty_mask = df["temps_site"].astype(str).str.strip().eq("")
    if empty_mask.any():
        df.loc[empty_mask, "temps_site"] = "30"
        log.info("    Valeur par dfaut 30 min applique  %d lignes", int(empty_mask.sum()))

    log.info("    Colonne 'temps_site' dtecte - Temps de visite personnalis activ")
    log.info("  📅 Colonnes detectees: %s", list(df.columns))

    raw_rows = df.to_dict("records")
    log.info("  ✅ %d lignes lues", len(raw_rows))

    # 2. Geocodage
    log.info("=== 2/7 Geocodage avec validation ===")
    _prog(10, f"Verification des coordonnees ({len(raw_rows)} sites)...")

    # Detection si donnees pre-geocodees
    is_pre_geocoded = any(
        pd.notna(r.get("_lat")) and pd.notna(r.get("_lon")) 
        for r in raw_rows
    )
    log.info(f"   Donnees pre-geocodees detectees : {is_pre_geocoded}")

    if is_pre_geocoded:
        # CAS 1 : Donnees deja geocodees (vient de l'etape validation)
        all_geocoded_sites = []
        for r in raw_rows:
            # Conversion string vers float pour eviter les erreurs de logs/calculs
            try:
                lat = float(str(r.get("_lat", "")).replace(",", "."))
                lon = float(str(r.get("_lon", "")).replace(",", "."))
                if math.isfinite(lat) and math.isfinite(lon):
                    r["_lat"] = lat
                    r["_lon"] = lon
                    all_geocoded_sites.append(r)
            except (ValueError, TypeError):
                continue
                
        log.info(f"  ✅ {len(all_geocoded_sites)} sites utilises (coordonnees existantes)")
        
        # Verification presence Depot et ajout manuel si manquant
        depot_exists = _find_depot_index(all_geocoded_sites, depot_name) is not None
        
        if not depot_exists:
            log.warning(f"  🚚 Le depot '{depot_name}' n'est pas dans le fichier. Ajout manuel depuis les donnees brutes...")
            raw_depot_idx = _find_depot_index(raw_rows, depot_name)
            if raw_depot_idx is not None:
                r = raw_rows[raw_depot_idx]
                added = False
                try:
                    lat = float(str(r.get("_lat", "")).replace(",", "."))
                    lon = float(str(r.get("_lon", "")).replace(",", "."))
                    if not math.isfinite(lat) or not math.isfinite(lon):
                        raise ValueError("Coordonnees non finies")
                    r["_lat"] = lat
                    r["_lon"] = lon
                    r["_date_min"] = ""
                    r["_date_max"] = ""
                    all_geocoded_sites.insert(0, r)
                    log.info(f"  ✅ Depot ajoute manuellement en position 0 (lat={lat:.5f}, lon={lon:.5f})")
                    added = True
                except (ValueError, TypeError):
                    pass
                if not added:
                    # Fallback : geocoder l'adresse du depot
                    depot_addr  = str(r.get("address", "")).strip()
                    depot_ville_str = str(r.get("ville", "")).strip()
                    query = ", ".join(filter(None, [depot_addr, depot_ville_str]))
                    if query:
                        log.warning(f"   Geocodage du depot '{depot_name}': {query}")
                        geo = _geocode_one(query)
                        r["_lat"] = geo["lat"]
                        r["_lon"] = geo["lon"]
                        r["_date_min"] = ""
                        r["_date_max"] = ""
                        all_geocoded_sites.insert(0, r)
                        log.info(f"  ✅ Depot geocode et ajoute en position 0 (lat={geo['lat']:.5f}, lon={geo['lon']:.5f})")
                    else:
                        log.error(f"   Depot '{depot_name}' sans coordonnees ni adresse valide")
        
        sites = all_geocoded_sites
        
    else:
        # CAS 2 : Geocodage API
        geocode_results = geocode_with_validation(raw_rows, session_id)
        if geocode_results["manual_needed"] > 0:
            return {
                "success": False,
                "stage": "geocoding",
                "needs_manual": True,
                "message": f"{geocode_results['manual_needed']} adresses necessitent une correction manuelle",
                "geocode_results": geocode_results,
                "session_id": session_id
            }
        sites = geocode_results["success"]

    _prog(55, "Coordonnees chargees - analyse des dates...")

    # 2.5 Gestion des Dates (Metadonnees uniquement - Pas de contrainte VRP)
    # C'est ici que l'on met votre bloc de code
    log.info("=== 2.5/7 Analyse des dates (Metadonnees) ===")
    
    for s in sites:
        d_min = parse_date(s.get("date_min", ""))
        d_max = parse_date(s.get("date_max", ""))
        
        if d_min and d_max:
            s["_date_min"] = d_min.strftime("%d/%m/%Y")
            s["_date_max"] = d_max.strftime("%d/%m/%Y")
        else:
            s["_date_min"] = ""
            s["_date_max"] = ""
            
    log.info(f"  ✅ Metadonnees de dates ajoutees a {len(sites)} sites")

    # 3. Depot
    log.info("=== 3/7 Depot ===")
    depot_idx = _find_depot_index(sites, depot_name)
    if depot_idx is None:
        raise ValueError(f"Depot '{depot_name}' introuvable.")
    
    depot_site = sites[depot_idx]
    log.info("  Depot idx=%d  lat=%.5f lon=%.5f", depot_idx,
             depot_site["_lat"], depot_site["_lon"])

    # 4. Matrice
    log.info("=== 4/7 Matrice OSRM (%s) ===", profil)
    _prog(60, f"Calcul de la matrice de distances ({len(sites)} sites)...")
    time_m, dist_m = build_matrices(sites, profil)

    BIG = 10_000_000
    n = len(sites)
    non_big = sum(1 for i in range(n) for j in range(n) if i != j and time_m[i][j] < BIG)
    log.info("  Paires renseignees : %d / %d", non_big, n*(n-1))

        # 5. VRP Global (Base sur le temps)
    log.info("=== 5/7 VRP OR-Tools (Base sur le temps) ===")
    _prog(70, "Optimisation des tournees (OR-Tools)...")

    # Nouveau parametre: temps max par tournee en heures
    max_tour_hours = float(params.get("max_tour_hours", 7.0))
    log.info(f"  Contrainte: Max {max_tour_hours}h par tournee")

    if split_mode == "auto":
        # Mode auto : le VRP cre un seul grand parcours (TSP) sans contrainte
        # de temps, puis split_long_tours_by_time le dcoupe en tournes
        # de dure variable selon max_tour_hours + dure de visite.
        log.info("   Mode auto : rsolution TSP (1 vhicule sans contrainte de temps)")
        raw_routes = run_vrp_robust(time_m, depot_idx, 999, solver_time, sites)
    else:
        raw_routes = run_vrp_robust(time_m, depot_idx, max_tour_hours, solver_time, sites)

    if not raw_routes:
        raise RuntimeError("OR-Tools : aucune solution trouvee.")

    log.info(f"   VRP Brut : {len(raw_routes)} tournees trouvees.")

    # Decoupage des routes trop longues (base sur le temps incluant les visites)
    split_routes_indices = split_long_tours_by_time(
        raw_routes, time_m, depot_idx, max_s, split_size, sites,
        split_mode=split_mode, dist_m=dist_m, use_ai=use_ai
    )
    
    # Conversion finale en distributions
    routes = []
    for r in split_routes_indices:
        routes.append({
            "route": r,
            "group_idx": 0,
            "date_min": None,
            "date_max": None
        })
    
    log.info(f"  ✅ {len(routes)} tournees finales.")


    # 6. Analyse IA
    ai_analysis = None
    if use_ai:
        log.info("=== 6/7 Analyse IA ===")
        _prog(85, "Analyse IA des tournees...")
        ai_result = analyze_routes_with_ai(routes, sites, time_m, dist_m, depot_idx)
        if ai_result["success"]:
            ai_analysis = ai_result
            log.info("  ✅ Analyse IA terminee")
        else:
            log.warning("    Analyse IA echouee")
            ai_analysis = { "success": False, "message": "Erreur analyse" }

    # 7. Resultats
    log.info("=== 7/7 Resultats ===")
    _prog(93, "Generation du fichier Excel...")

    def route_totals(route):
        """
        Calcule le temps total incluant:
        - Temps de trajet
        - Temps de visite sur chaque site
        """
        # Temps de trajet : Depot -> Route -> Depot
        seq = [depot_idx] + route + [depot_idx]
        travel_time = sum(time_m[seq[k]][seq[k+1]] for k in range(len(seq)-1))
        travel_dist = sum(dist_m[seq[k]][seq[k+1]] for k in range(len(seq)-1))
        
        # Temps de visite sur chaque site
        visit_time = 0
        for idx in route:
            visit_time += _visit_minutes_from_row(sites[idx], default=30) * 60
        
        return travel_time + visit_time, travel_dist, visit_time

    tours_json  = []
    tours_excel = []
    total_km = total_h = total_visits = 0
    issues = []

    for tour_num, route_data in enumerate(routes, 1):
        route = route_data["route"]
        
        ts, dm, visit_time = route_totals(route)
        total_km     += dm / 1000
        total_h      += ts / 3600
        total_visits += len(route)

        # Detection de problemes
        if ts > max_s * 0.9:
            issues.append({
                "tour": tour_num,
                "type": "long_tour",
                "message": f"Tournee #{tour_num} proche de la duree maximale ({ts/3600:.1f}h / {max_s/3600:.1f}h)"
            })
        
        # Log detaille par tournee
        log.info(f"   T{tour_num}: {len(route)} sites, {dm/1000:.1f}km, {ts/3600:.1f}h (trajet: {(ts-visit_time)/3600:.1f}h + visite: {visit_time/3600:.1f}h)")

        # Construction de la route pour la carte et le JSON
        route_coords = [[depot_site["_lat"], depot_site["_lon"]]]
        stops = []
        
        # IMPORTANT : On filtre le depot pour ne pas l'afficher comme un site
        for si in route:
            s = sites[si]
            # Double verification : si c'est le depot, on passe
            if si != depot_idx:
                stops.append({
                    "site":    str(s.get("site",   "")),
                    "client":  str(s.get("client", "")),
                    "ville":   str(s.get("ville",  "")),
                    "address": str(s.get("address","")),
                    "lat":     s.get("_lat"),
                    "lon":     s.get("_lon"),
                    "date_min": str(s.get("_date_min", "")),
                    "date_max": str(s.get("_date_max", "")),
                })
                route_coords.append([s["_lat"], s["_lon"]])
        
        # On revient au depot pour fermer la boucle visuelle
        route_coords.append([depot_site["_lat"], depot_site["_lon"]])

        # Calcul de la fenetre de dates valide pour cette tournee
        # (intersection des fenetres de tous les sites de la tournee)
        tour_dates_min, tour_dates_max = [], []
        for stop in stops:
            d = parse_date(stop.get("date_min", ""))
            if d: tour_dates_min.append(d)
            d = parse_date(stop.get("date_max", ""))
            if d: tour_dates_max.append(d)

        tour_date_min_str = None
        tour_date_max_str = None
        tour_date_proposed = "A definir"

        if tour_dates_min and tour_dates_max:
            t_min = max(tour_dates_min)   # plus tardive des dates min
            t_max = min(tour_dates_max)   # plus precoce des dates max
            if t_min <= t_max:
                tour_date_min_str = t_min.strftime("%d/%m/%Y")
                tour_date_max_str = t_max.strftime("%d/%m/%Y")
                # Premier jour ouvrable (lun-ven) dans la fenetre
                d = t_min
                while d <= t_max:
                    if d.weekday() < 5:
                        tour_date_proposed = d.strftime("%d/%m/%Y")
                        break
                    d += timedelta(days=1)
                else:
                    tour_date_proposed = t_min.strftime("%d/%m/%Y")
            else:
                tour_date_proposed = " Conflit de dates"
                issues.append({
                    "tour": tour_num,
                    "type": "date_conflict",
                    "message": f"Tournee #{tour_num} : les fenetres de dates des sites sont incompatibles"
                })
        elif tour_dates_min:
            tour_date_min_str = max(tour_dates_min).strftime("%d/%m/%Y")
        elif tour_dates_max:
            tour_date_max_str = min(tour_dates_max).strftime("%d/%m/%Y")

        tour_observation = "Tournee optimisee."

        tours_json.append({
            "tour_num":          tour_num,
            "group_idx":         0,
            "date_min":          tour_date_min_str,
            "date_max":          tour_date_max_str,
            "date_proposed":     tour_date_proposed,
            "observation":       tour_observation,
            "total_distance_km": round(dm/1000, 2),
            "total_time_hours":  round(ts/3600, 2),
            "travel_time_hours": round((ts - visit_time) / 3600, 2),
            "visit_time_hours": round(visit_time / 3600, 2),
            "stops":             stops,
            "route_coords":      route_coords,
        })
        tours_excel.append({
            "tour_num": tour_num,
            "group_idx": 0,
            "total_km": round(dm/1000, 2),
            "total_h":  round(ts/3600, 2),
            "stops":    stops,
        })

    excel_b64 = build_excel(tours_excel)

    return {
        "success": True,
        "stats": {
            "total_tours":       len(tours_json),
            "total_visits":      total_visits,
            "total_distance_km": round(total_km, 2),
            "total_time_hours":  round(total_h,  2),
            "geocode_warnings":  0,
            "issues_detected":   len(issues)
        },
        "tours": tours_json,
        "depot": {
            "lat":     depot_site["_lat"],
            "lon":     depot_site["_lon"],
            "address": str(depot_site.get("address", "")),
        },
        "date_groups": [],
        "output_file_base64": excel_b64,
        "ai_analysis": ai_analysis,
        "issues": issues,
        "session_id": session_id
    }




# ─────────────────────────────────────────────────────────────────
# ROUTES API
# ─────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return send_from_directory(".", "index.html")

@app.route("/api/upload-validate", methods=["POST"])
def api_upload_validate():
    """
    Etape 1 : Upload et validation/geocodage des adresses
    Retourne les adresses validees et celles necessitant une correction
    """
    try:
        if "file" not in request.files:
            return jsonify({"success": False, "error": "Aucun fichier uploade"}), 400
        f = request.files["file"]
        if not f.filename:
            return jsonify({"success": False, "error": "Nom de fichier vide"}), 400

        params = {
            "depot_site": request.form.get("depot_site", "depot_ST_FOY"),
            "profil":     request.form.get("profil", "driving"),
        }

        progress_id   = request.form.get("progress_id", "")
        mapping_raw   = request.form.get("column_mapping", "")
        column_mapping = json.loads(mapping_raw) if mapping_raw else {}

        ts    = datetime.now().strftime("%Y%m%d_%H%M%S")
        session_id = f"{ts}_{secure_filename(f.filename)}"
        fname = f"{session_id}.xlsx"
        fpath = os.path.join(app.config["UPLOAD_FOLDER"], fname)
        f.save(fpath)

        log.info("Upload pour validation : %s (mapping=%s)", session_id, list(column_mapping.keys()))

        try:
            set_progress(progress_id, 5, "Lecture du fichier...")
            df = pd.read_excel(fpath, dtype=str, na_filter=False, keep_default_na=False)
            if df.empty:
                raise ValueError("Le fichier Excel est vide.")

            # Appliquer le mappage de colonnes recu du client
            if column_mapping:
                df = apply_column_mapping(df, column_mapping)
                log.info("  📌 Mappage applique : %s", column_mapping)

            # Colonnes obligatoires manquantes  valeur vide
            for col in ["site", "address", "ville"]:
                if col not in df.columns:
                    df[col] = ""

            # Colonnes optionnelles : si absentes, creer vides
            for col in ["date_min", "date_max", "temps_site"]:
                if col not in df.columns:
                    df[col] = ""

            # Harmonisation du temps de visite :
            # - 'Temps' est en heures dans le fichier metier
            # - 'temps_site' est en minutes dans l'application
            if "Temps" in df.columns:
                def _h_to_min(v):
                    try:
                        h = float(str(v).strip().replace(",", "."))
                        return str(int(round(h * 60))) if h > 0 else ""
                    except (ValueError, TypeError):
                        return ""
                hours_mask = df["Temps"].astype(str).str.strip().ne("")
                if hours_mask.any():
                    df.loc[hours_mask, "temps_site"] = df.loc[hours_mask, "Temps"].apply(_h_to_min)
                    log.info("    Upload: 'Temps' (heures) converti vers 'temps_site' (minutes) sur %d lignes", int(hours_mask.sum()))

            raw_rows = df.to_dict("records")
            set_progress(progress_id, 10, f"Geocodage de {len(raw_rows)} adresses...")

            def _geo_cb(pct, msg):
                set_progress(progress_id, pct, msg)

            geocode_results = geocode_with_validation(raw_rows, session_id, progress_callback=_geo_cb)
            set_progress(progress_id, 100, "Geocodage termine ✓", done=True)

            return jsonify({
                "success": True,
                "session_id": session_id,
                "geocode_results": geocode_results,
                "depot_site": params["depot_site"]
            })
            
        finally:
            try: os.remove(fpath)
            except OSError: pass

    except Exception as e:
        log.exception("Erreur API /api/upload-validate")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/manual-geocode", methods=["POST"])
def api_manual_geocode():
    """
    Permet de corriger manuellement les adresses et de les geocoder
    """
    try:
        data = request.json
        session_id = data.get("session_id")
        corrections = data.get("corrections", [])
        
        log.info(f"Correction manuelle demandee - Session: {session_id}, Corrections: {len(corrections)}")
        
        if not session_id or session_id not in temp_data_store:
            log.error("Session invalide: %s", session_id)
            return jsonify({"success": False, "error": "Session invalide"}), 400
        
        results = temp_data_store[session_id]
        
        # Traiter les corrections
        corrected_count = 0
        for corr in corrections:
            idx = corr.get("index")
            new_address = corr.get("new_address", "")
            lat = corr.get("lat")
            lon = corr.get("lon")
            
            log.info(f"Traitement correction - Index: {idx}, Adresse: {new_address}")
            
            if idx is not None and (new_address or (lat and lon)):
                # Retrouver l'entree dans failed par son index original
                failed_entry = next(
                    (f for f in results["failed"] if f.get("_original_index") == idx),
                    None
                )
                
                if failed_entry:
                    # Priorite aux coordonnees directes
                    if lat and lon:
                        try:
                            failed_entry["_lat"] = float(lat)
                            failed_entry["_lon"] = float(lon)
                            failed_entry["_display_name"] = f"Coordonnees manuelles: {lat}, {lon}"
                            failed_entry["_normalized_address"] = new_address or failed_entry.get("_normalized_address", "")
                            failed_entry["_manually_corrected"] = True
                            failed_entry["_needs_manual"] = False
                            failed_entry["_coords_manual"] = True
                            log.info("✅ Coordonnees manuelles appliquees pour index %d", idx)
                        except (TypeError, ValueError) as e:
                            log.error(" Erreur coordonnees: %s", e)
                            continue
                    # Sinon, geocodage avec nouvelle adresse
                    elif new_address:
                        geo_result = _geocode_one(new_address)
                        
                        # _geocode_one ne retourne PLUS JAMAIS None
                        if geo_result:
                            failed_entry.update({
                                "_lat": geo_result["lat"],
                                "_lon": geo_result["lon"],
                                "_display_name": geo_result["display_name"],
                                "_normalized_address": new_address,
                                "_manually_corrected": True,
                                "_needs_manual": False,
                                "_is_fallback": geo_result.get("is_fallback", False)
                            })
                            log.info("✅ Geocodage reussi pour index %d", idx)
                        else:
                            # Ceci ne devrait plus jamais arriver
                            log.warning("  Geocodage echoue pour index %d (impossible)", idx)
                            continue
                    
                    # Deplacer vers success
                    results["success"].append(failed_entry)
                    results["failed"] = [f for f in results["failed"] if f.get("_original_index") != idx]
                    results["geocoded"] += 1
                    results["manual_needed"] -= 1
                    corrected_count += 1
                else:
                    log.warning("  Entree non trouvee pour index %d", idx)
        
        # Mettre a jour le stockage
        temp_data_store[session_id] = results
        
        log.info("✅ Correction terminee: %d adresses corrigees", corrected_count)
        
        return jsonify({
            "success": True,
            "geocode_results": results,
            "corrected_count": corrected_count
        })
        
    except Exception as e:
        log.exception("Erreur API /api/manual-geocode")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/optimize", methods=["POST"])
def api_optimize():
    """
    Etape 2 : Optimisation complete avec les adresses validees
    """
    try:
        # On essaie de lire le JSON
        try:
            data = request.json
        except:
            data = None
        
        progress_id = data.get("progress_id", "") if data else ""

        if not data or "session_id" not in data:
            # --- Fallback pour l'ancienne methode avec upload de fichier ---
            if "file" not in request.files:
                return jsonify({"success": False, "error": "Aucun fichier ni session_id"}), 400

            f = request.files["file"]
            session_id = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{secure_filename(f.filename)}"
            fname = f"{session_id}.xlsx"
            fpath = os.path.join(app.config["UPLOAD_FOLDER"], fname)
            f.save(fpath)

            params = {
                "depot_site":     request.form.get("depot_site",    "depot_ST_FOY"),
                "max_tour_hours": float(request.form.get("max_tour_hours", 7.0)),
                "solver_time":    int(request.form.get("solver_time",  45)),
                "split_size":     int(request.form.get("split_size",   3)),
                "split_mode":     request.form.get("split_mode", "manual"),
                "profil":         request.form.get("profil", "driving"),
                "use_ai":         request.form.get("use_ai", "true").lower() == "true",
            }

            try:
                result = full_pipeline_enhanced(fpath, params, session_id, progress_id=progress_id)
            finally:
                try: os.remove(fpath)
                except OSError: pass
        else:
            # --- Nouvelle methode avec session_id ---
            session_id = data["session_id"]
            params = {
                "depot_site":     data.get("depot_site",    "depot_ST_FOY"),
                "max_tour_hours": float(data.get("max_tour_hours", 7.0)),
                "solver_time":    int(data.get("solver_time",  45)),
                "split_size":     int(data.get("split_size",   3)),
                "split_mode":     data.get("split_mode", "manual"),
                "profil":         data.get("profil", "driving"),
                "use_ai":         data.get("use_ai", True),
            }

            # Recuperer les resultats de geocodage
            if session_id not in temp_data_store:
                return jsonify({"success": False, "error": "Session expiree"}), 400

            geocode_results = temp_data_store[session_id]

            if geocode_results["manual_needed"] > 0:
                return jsonify({
                    "success": False,
                    "stage": "geocoding",
                    "needs_manual": True,
                    "message": "Corrigez d'abord les adresses manuellement"
                })

            # Creer un fichier temporaire avec les sites geocodes
            sites = geocode_results["success"]

            # Creer un DataFrame temporaire
            df_temp = pd.DataFrame(sites)

            # Sauvegarder temporairement
            fpath = os.path.join(app.config["UPLOAD_FOLDER"], f"{session_id}_geocoded.xlsx")
            df_temp.to_excel(fpath, index=False)

            try:
                result = full_pipeline_enhanced(fpath, params, session_id, progress_id=progress_id)
            finally:
                try: os.remove(fpath)
                except OSError: pass

        set_progress(progress_id, 100, "Optimisation terminee ✓", done=True)
        return jsonify(result)

    except Exception as e:
        log.exception("Erreur API /api/optimize")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/move-stop", methods=["POST"])
def api_move_stop():
    """
    Deplace un stop d'une tournee a une autre ou reordonne dans la meme tournee.
    """
    try:
        data = request.json
        session_id = data.get("session_id")
        tours = data.get("tours", [])
        from_tour = data.get("from_tour")
        from_idx = data.get("from_idx")
        to_tour = data.get("to_tour")
        to_idx = data.get("to_idx")

        if not tours or from_tour is None or from_idx is None or to_tour is None or to_idx is None:
            return jsonify({"success": False, "error": "Parametres manquants"}), 400

        src = next((t for t in tours if t["tour_num"] == from_tour), None)
        dst = next((t for t in tours if t["tour_num"] == to_tour), None)

        if not src or not dst:
            return jsonify({"success": False, "error": "Tournee introuvable"}), 400

        if from_idx < 0 or from_idx >= len(src["stops"]):
            return jsonify({"success": False, "error": "Index stop invalide"}), 400

        stop = src["stops"].pop(from_idx)
        
        # Ajuster l'index si meme tournee et deplacement vers l'avant
        insert_idx = to_idx
        if from_tour == to_tour and to_idx > from_idx:
            insert_idx = max(0, to_idx - 1)
        
        dst["stops"].insert(insert_idx, stop)

        # Reconstruire route_coords
        def rebuild_coords(t, depot_info):
            coords = []
            if depot_info:
                coords.append([depot_info["lat"], depot_info["lon"]])
            for stp in t["stops"]:
                if stp.get("lat") and stp.get("lon"):
                    coords.append([stp["lat"], stp["lon"]])
            if depot_info:
                coords.append([depot_info["lat"], depot_info["lon"]])
            t["route_coords"] = coords

        depot_info = None
        if session_id and session_id in temp_data_store:
            geo_results = temp_data_store[session_id]
            success_sites = geo_results.get("success", [])
            di = _find_depot_index(success_sites, "depot_ST_FOY")
            if di is not None:
                s = success_sites[di]
                depot_info = {"lat": s["_lat"], "lon": s["_lon"]}

        rebuild_coords(src, depot_info)
        if from_tour != to_tour:
            rebuild_coords(dst, depot_info)

        return jsonify({
            "success": True,
            "tours": tours,
            "message": f"Stop deplace: T{from_tour}[{from_idx}]  T{to_tour}[{insert_idx}]"
        })

    except Exception as e:
        log.exception("Erreur API /api/move-stop")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/recalculate-tour", methods=["POST"])
def api_recalculate_tour():
    """
    Recalcule les temps/distances d'une tournee modifiee manuellement via OSRM.
    """
    try:
        data = request.json
        session_id = data.get("session_id")
        tour = data.get("tour")
        profil = data.get("profil", "driving")

        if not tour or not session_id:
            return jsonify({"success": False, "error": "Donnees manquantes"}), 400

        if session_id not in temp_data_store:
            return jsonify({"success": False, "error": "Session expiree"}), 400

        geo_results = temp_data_store[session_id]
        success_sites = geo_results.get("success", [])
        di = _find_depot_index(success_sites, "depot_ST_FOY")
        depot_site = success_sites[di] if di is not None else None

        if not depot_site:
            return jsonify({"success": False, "error": "Depot introuvable"}), 400

        stops = tour.get("stops", [])
        if not stops:
            tour.update({"total_distance_km": 0, "total_time_hours": 0,
                         "travel_time_hours": 0, "visit_time_hours": 0})
            return jsonify({"success": True, "tour": tour})

        sites_list = [depot_site] + [
            {
                "_lat": float(stp.get("lat", 0)),
                "_lon": float(stp.get("lon", 0)),
                "site": stp.get("site", ""),
                "temps_site": str(stp.get("temps_site_min", 30)),
            }
            for stp in stops
            if stp.get("lat") and stp.get("lon")
        ]

        if len(sites_list) < 2:
            return jsonify({"success": True, "tour": tour})

        time_m, dist_m = build_matrices(sites_list, profil)
        n = len(sites_list)
        seq = list(range(n))

        travel_time = sum(time_m[seq[k]][seq[k+1]] for k in range(len(seq)-1))
        travel_time += time_m[seq[-1]][0]
        travel_dist = sum(dist_m[seq[k]][seq[k+1]] for k in range(len(seq)-1))
        travel_dist += dist_m[seq[-1]][0]

        visit_time = 0
        for stp in stops:
            try:
                visit_time += int(float(stp.get("temps_site_min", 30))) * 60
            except (ValueError, TypeError):
                visit_time += 1800

        total_time = travel_time + visit_time

        route_coords = [[depot_site["_lat"], depot_site["_lon"]]]
        for stp in stops:
            if stp.get("lat") and stp.get("lon"):
                route_coords.append([float(stp["lat"]), float(stp["lon"])])
        route_coords.append([depot_site["_lat"], depot_site["_lon"]])

        tour.update({
            "total_distance_km": round(travel_dist / 1000, 2),
            "total_time_hours": round(total_time / 3600, 2),
            "travel_time_hours": round(travel_time / 3600, 2),
            "visit_time_hours": round(visit_time / 3600, 2),
            "route_coords": route_coords,
        })

        return jsonify({"success": True, "tour": tour})

    except Exception as e:
        log.exception("Erreur API /api/recalculate-tour")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/auto-split", methods=["POST"])
def api_auto_split():
    """
    Calcule automatiquement le split_size optimal (nb de sites par tournee)
    en fonction du temps de trajet reel (OSRM) + temps de visite des sites.

    Algorithme :
    1. Recupere les sites geocodes depuis la session
    2. Calcule la matrice OSRM (ou utilise une estimation heuristique si trop de sites)
    3. Simule des tournees type et determine combien de sites tiennent dans max_tour_hours
    4. Retourne split_size + explication detaillee
    """
    try:
        data = request.json
        session_id   = data.get("session_id")
        max_hours    = float(data.get("max_tour_hours", 7.0))
        profil       = data.get("profil", "driving")
        depot_name   = data.get("depot_site", "depot_ST_FOY")

        if not session_id or session_id not in temp_data_store:
            return jsonify({"success": False, "error": "Session introuvable"}), 400

        geo_results = temp_data_store[session_id]
        all_sites   = geo_results.get("success", [])

        # Filtrer le depot
        depot_idx = _find_depot_index(all_sites, depot_name)
        depot_site = all_sites[depot_idx] if depot_idx is not None else None
        client_sites = [s for i, s in enumerate(all_sites) if i != depot_idx]

        if not client_sites:
            return jsonify({"success": False, "error": "Aucun site client trouve"}), 400

        n_sites = len(client_sites)
        max_s   = max_hours * 3600  # secondes

        # ── Calcul du temps de visite moyen ─────────────────────
        visit_times = []
        for s in client_sites:
            vt = _visit_minutes_from_row(s, default=30) * 60
            visit_times.append(vt)

        avg_visit_s  = sum(visit_times) / max(1, len(visit_times))
        total_visit_s = sum(visit_times)

        # ── Estimation du temps de trajet inter-sites ────────────
        # On calcule la matrice OSRM sur un echantillon (max 20 sites)
        # pour avoir un temps de trajet moyen realiste
        SAMPLE_SIZE = min(20, n_sites)
        sample_sites = client_sites[:SAMPLE_SIZE]
        if depot_site:
            sample_with_depot = [depot_site] + sample_sites
        else:
            sample_with_depot = sample_sites

        log.info("🔢 Auto-split: calcul OSRM sur %d sites (echantillon)", len(sample_with_depot))

        try:
            time_m_sample, dist_m_sample = build_matrices(sample_with_depot, profil)
            depot_i = 0 if depot_site else None

            # Temps de trajet moyen entre deux sites consecutifs dans une tournee optimale
            # On mesure le temps moyen de tous les arcs non-nuls
            arcs = []
            n_s = len(sample_with_depot)
            for i in range(n_s):
                for j in range(n_s):
                    if i != j and time_m_sample[i][j] < 5_000_000:
                        arcs.append(time_m_sample[i][j])

            if arcs:
                avg_travel_s = sorted(arcs)[len(arcs) // 3]  # Percentile 33% = trajet typique proche
            else:
                avg_travel_s = 1800  # 30 min fallback

            # Trajet depot  premier site et dernier site  depot
            if depot_site and len(sample_with_depot) > 1:
                depot_arcs = [time_m_sample[0][j] for j in range(1, n_s) if time_m_sample[0][j] < 5_000_000]
                avg_depot_travel_s = (sum(depot_arcs) / max(1, len(depot_arcs))) if depot_arcs else 3600
            else:
                avg_depot_travel_s = 3600  # 1h aller-retour fallback

            use_osrm = True
            log.info("✅ OSRM OK - trajet moyen inter-sites: %.0fs (%.1f min)", avg_travel_s, avg_travel_s/60)

        except Exception as e:
            log.warning("  OSRM echec (%s), estimation heuristique", e)
            avg_travel_s       = 1800   # 30 min entre sites
            avg_depot_travel_s = 3600   # 1h aller-retour depot
            use_osrm = False

        # ── Simulation : combien de sites tiennent dans max_s ? ──
        # On simule une tournee "type" :
        # temps_total = depots1 + (n-1)*avg_inter + sndepot + sum(visites)
        # On cherche le plus grand n tel que temps_total  max_s

        best_n = 1
        for n in range(1, n_sites + 1):
            # Temps de trajet estime pour n sites
            if n == 1:
                travel = avg_depot_travel_s  # aller-retour direct
            else:
                travel = avg_depot_travel_s + (n - 1) * avg_travel_s

            # Temps de visite : moyenne ponderee sur les n premiers sites
            visit_sample = visit_times[:n] if n <= len(visit_times) else visit_times
            visit = sum(visit_sample[:n]) if n <= len(visit_sample) else sum(visit_sample) * (n / max(1, len(visit_sample)))

            total = travel + visit

            if total <= max_s:
                best_n = n
            else:
                break

        # Garde-fous Auto:
        # - minimum 2 sites/tournee (si possible)
        # - maximum 8 sites/tournee
        if n_sites >= 2:
            best_n = max(2, min(best_n, 8, n_sites))
        else:
            best_n = max(1, min(best_n, n_sites))

        # ── Calcul du nombre de tournees estime ─────────────────
        estimated_tours = math.ceil(n_sites / best_n)

        # ── Detail pedagogique ───────────────────────────────────
        travel_for_best = avg_depot_travel_s + max(0, best_n - 1) * avg_travel_s
        visit_for_best  = sum(visit_times[:best_n]) if best_n <= len(visit_times) else avg_visit_s * best_n
        total_for_best  = travel_for_best + visit_for_best
        margin_s        = max_s - total_for_best

        detail = (
            f"{n_sites} sites  visite moy. {avg_visit_s/60:.0f} min/site  "
            f"trajet moy. {avg_travel_s/60:.0f} min/inter  "
            f"aller-retour depot ~{avg_depot_travel_s/60:.0f} min  "
            f"{best_n} sites/tournee ({total_for_best/3600:.1f}h/{max_hours}h, marge {margin_s/60:.0f} min)  "
            f"~{estimated_tours} tournees estimees"
        )

        log.info("✅ Auto-split calcule: %d sites/tournee (%s)", best_n, "OSRM" if use_osrm else "heuristique")

        return jsonify({
            "success":         True,
            "split_size":      best_n,
            "split_mode":      "auto",
            "estimated_tours": estimated_tours,
            "detail":          detail,
            "stats": {
                "n_sites":            n_sites,
                "avg_visit_min":      round(avg_visit_s / 60, 1),
                "avg_travel_min":     round(avg_travel_s / 60, 1),
                "avg_depot_min":      round(avg_depot_travel_s / 60, 1),
                "total_for_best_h":   round(total_for_best / 3600, 2),
                "margin_min":         round(margin_s / 60, 1),
                "used_osrm":          use_osrm,
            }
        })

    except Exception as e:
        log.exception("Erreur API /api/auto-split")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/test-groq", methods=["GET"])
def api_test_groq():
    """Test la connexion a Groq"""
    try:
        response = ask_groq("Test. Reponds juste OK.", timeout=15)
        if response and len(response) > 0:
            # Verifier que la reponse n'est pas vide
            return jsonify({
                "success": True,
                "message": "Groq connecte et fonctionnel",
                "model": GROQ_MODEL,
                "response_preview": response[:100]  # Apercu pour debug
            })
        else:
            return jsonify({
                "success": False,
                "message": "Groq connecte mais ne repond pas",
                "response": response
            })
    except requests.exceptions.Timeout:
        return jsonify({
            "success": False,
            "message": "Timeout - Groq trop lent"
        }), 200
    except requests.exceptions.ConnectionError:
        return jsonify({
            "success": False,
            "message": "Groq non connecte - Verifiez qu'il tourne"
        }), 200
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Erreur: {str(e)}"
        }), 200


@app.route("/health")
def health():
    return jsonify({
        "status": "ok", 
        "service": "VRP Optimizer Pro v4",
        "ai_provider": "groq",
        "model": GROQ_MODEL
    })

if __name__ == "__main__":
    print("=" * 70)
    print("  🚚  VRP Optimizer Pro v4")
    print("  Interface : http://localhost:5000")
    print("  API       : http://localhost:5000/api/optimize")
    print("  AI (Groq) : %s (cle %s)" % (GROQ_MODEL, "configuree" if GROQ_API_KEY else "MANQUANTE"))
    print("=" * 70)
    app.run(host="0.0.0.0", port=5000, debug=False, threaded=True)

